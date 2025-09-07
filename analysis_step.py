import csv
import inspect
import math
import os
import re

import h5py
import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from matplotlib.backends.backend_qt5agg import (
    FigureCanvasQTAgg as FigureCanvas,
)
from matplotlib.figure import Figure
from matplotlib.path import Path
from matplotlib.widgets import LassoSelector, RectangleSelector
from PyQt5.QtCore import QObject, Qt, QThread, QTimer, pyqtSignal
from PyQt5.QtWidgets import (
    QApplication,
    QButtonGroup,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QGridLayout,
    QHBoxLayout,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMenu,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QSizePolicy,
    QSlider,
    QTableWidget,
    QTableWidgetItem,
    QTextEdit,
    QVBoxLayout,
    QWidget,
)

mpl.use("Qt5Agg")
mpl.rcParams["agg.path.chunksize"] = 20000  # speed up long paths
mpl.rcParams["path.simplify"] = True
mpl.rcParams["path.simplify_threshold"] = 0.3

plt.rcParams["font.family"] = "Times New Roman"


def _make_lasso_selector(ax, onselect):
    """Create a LassoSelector with a GREEN line across Matplotlib
    versions."""
    props = {"color": "green", "linewidth": 1.5, "alpha": 0.9}
    sig = inspect.signature(LassoSelector.__init__)
    if "props" in sig.parameters:
        return LassoSelector(ax=ax, onselect=onselect, useblit=True, props=props)
    else:
        return LassoSelector(ax=ax, onselect=onselect, useblit=True, lineprops=props)


def _safe_fisher_ci(r, n, alpha=0.05):
    if not np.isfinite(r) or n < 4:
        return (np.nan, np.nan)
    z = np.arctanh(np.clip(r, -0.999999, 0.999999))
    se = 1.0 / np.sqrt(max(n - 3, 1))
    z_lo = z + np.sqrt(2) * se * np.erfinv(2 * alpha / 2 - 1)  # ~-1.96*se for 95%
    z_hi = z - np.sqrt(2) * se * np.erfinv(2 * alpha / 2 - 1)  # ~+1.96*se
    return (np.tanh(z_lo), np.tanh(z_hi))


def _weighted_pearson(x, y, w):
    m = np.isfinite(x) & np.isfinite(y) & np.isfinite(w) & (w > 0)
    if not np.any(m):
        return np.nan
    x = x[m]
    y = y[m]
    w = w[m]
    w = np.sum(w)
    mx = np.sum(w * x) / w
    my = np.sum(w * y) / w
    cov = np.sum(w * (x - mx) * (y - my)) / w
    vx = np.sum(w * (x - mx) ** 2) / w
    vy = np.sum(w * (y - my) ** 2) / w
    denom = np.sqrt(vx * vy)
    return cov / denom if denom > 0 else np.nan


def _estimate_sigma_from_counts(arr, scale=1.0):
    a = np.asarray(arr, dtype=float)
    a = np.clip(a, 0, None)
    return scale * np.sqrt(a)


def _debounce(self, attr, fn, ms=40):
    """Call fn() once after ms; collapse bursts.

    attr is a string timer name.
    """
    t = getattr(self, attr, None)
    if t is None:
        t = QTimer(self)
        t.setSingleShot(True)
        setattr(self, attr, t)
        t.timeout.connect(fn)
    t.start(ms)


# -----------------------------
# Data loading thread
# -----------------------------
class DataLoader(QThread):
    """DataLoader class."""

    data_loaded = pyqtSignal(list, list)

    def __init__(self, data_path=None):
        super().__init__()
        self.working_dir = os.environ.get("MAPEX_WORKDIR", "")
        self.data_path = data_path or "data.h5"
        if not os.path.isabs(self.data_path):
            # resolve relative paths into the working folder chosen in main.py
            self.data_path = os.path.abspath(
                os.path.join(
                    self.working_dir or os.getcwd(),
                    os.path.basename(self.data_path),
                )
            )

    def run(self):
        """Run."""
        try:
            file_path = self.data_path or "data.h5"
            data_list, image_list = [], []
            with h5py.File(file_path, "r") as f:
                if "X-ray Maps" in f:
                    for group_name, group in f["X-ray Maps"].items():
                        if "data" in group:
                            data_list.append((group_name, group["data"][()]))
                if "Images" in f:
                    for image_name, dataset in f["Images"].items():
                        image_list.append((image_name, dataset[()]))
            self.data_loaded.emit(data_list, image_list)
        except Exception:
            self.data_loaded.emit([], [])

    """SelectionManager class.

    """


class SelectionManager(QObject):
    selection_changed = pyqtSignal(set)  # emits set of flat indices

    def __init__(self, parent=None):
        super().__init__(parent)
        """Clear."""
        self._indices = set()

        """Set.

        """

    def clear(self, quiet=False):
        """Add."""
        self._indices.clear()
        if not quiet:
            self.selection_changed.emit(set())
        """Remove.

        """

    def set(self, indices):
        self._indices = set(indices) if indices else set()
        self.selection_changed.emit(set(self._indices))
        """Has any."""
        """Get."""

    def add(self, indices):
        before = set(self._indices)
        self._indices |= set(indices)
        if self._indices != before:
            self.selection_changed.emit(set(self._indices))

    def remove(self, indices):
        before = set(self._indices)
        for i in indices:
            self._indices.discard(i)
        if self._indices != before:
            self.selection_changed.emit(set(self._indices))

    def has_any(self):
        return bool(self._indices)

    def get(self):
        return set(self._indices)

    """BasePlotWidget class.

    """


# -----------------------------
# Base plotting widget
# -----------------------------
class BasePlotWidget(FigureCanvas):
    def __init__(self, width=4, height=3, dpi=100):
        self.figure = Figure(figsize=(width, height), dpi=dpi)
        self.ax = self.figure.add_subplot(111)
        super().__init__(self.figure)
        self.setFocusPolicy(Qt.ClickFocus)
        self.setFocus()
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        try:
            self.figure.set_tight_layout(True)
        except Exception:
            pass

        # -----------------------------
        # Element map panel (single channel)
        """Set data.

        """


# -----------------------------
class ElementMapView(BasePlotWidget):
    roi_selected = pyqtSignal(tuple)  # (x1, y1, x2, y2)
    line_selected = pyqtSignal(tuple)  # (x1, y1, x2, y2)

    def __init__(self, title=""):
        super().__init__()
        self.title = None
        self.data = None
        self.im = None
        self.rect_selector = None
        self.line_start = None
        self._connect_mouse()
        self.ax.set_title(self.title)
        self.offset = (0, 0)
        self._hl_scatter = None
        self._last_highlight = []
        self._cursor_mk = None
        self._hl_visible = True  # visibility toggle

    def set_data(self, data, cmap="Reds", title=None, offset=(0, 0)):
        self.data = data
        self.offset = offset
        self.ax.clear()
        """Set highlight visible."""
        if data is None or data.size == 0:
            self.ax.set_title(title or self.title or "")
            _debounce(self, "_draw_timer", self.draw_idle, ms=40)
            return
        self.im = self.ax.imshow(data, cmap=cmap)
        self.ax.set_title(title or self.title)
        h, w = data.shape[:2]
        if w > 0 and h > 0:
            self.ax.set_xlim(-0.5, w - 0.5)
            self.ax.set_ylim(h - 0.5, -0.5)
        if getattr(self, "_last_highlight", None):
            self._ensure_hl_artist()
            self._update_hl_artist()
        self._ensure_cursor()
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)

    # ---------- Highlight overlay ----------
    def _ensure_hl_artist(self):
        if self._hl_scatter is None or getattr(self._hl_scatter, "axes", None) is None:
            self._hl_scatter = self.ax.scatter(
                [],
                [],
                s=18,
                facecolors="none",
                edgecolors="yellow",
                linewidths=1.0,
                zorder=6,
            )
        """Set highlight points.

        """

    def set_highlight_visible(self, flag: bool):
        self._hl_visible = bool(flag)
        if self._hl_scatter is not None:
            self._hl_scatter.set_visible(self._hl_visible and len(self._last_highlight) > 0)
            _debounce(self, "_draw_timer", self.draw_idle, ms=40)

    def _update_hl_artist(self):
        if self.data is None:
            return
        self._ensure_hl_artist()
        ox, oy = self.offset
        h, w = self.data.shape[:2]
        xs, ys = [], []
        for x, y in self._last_highlight or []:
            lx, ly = x - ox, y - oy
            if 0 <= lx < w and 0 <= ly < h:
                xs.append(lx)
                ys.append(ly)
        if xs:
            self._hl_scatter.set_offsets(np.c_[xs, ys])
            self._hl_scatter.set_visible(self._hl_visible)
        else:
            self._hl_scatter.set_offsets(np.empty((0, 2)))
            self._hl_scatter.set_visible(False)
        """Set roi enabled.

        """

    def set_highlight_points(self, coords_xy_list):
        self._last_highlight = coords_xy_list or []
        self._update_hl_artist()
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)

    # ---------- Live cursor arrow ----------
    def _ensure_cursor(self):
        if self._cursor_mk is None or getattr(self._cursor_mk, "axes", None) is None:
            (self._cursor_mk,) = self.ax.plot([], [], marker="^", linestyle="None", markersize=6, zorder=7)

    def set_cursor_point(self, x, y):
        if self.data is None:
            return
        self._ensure_cursor()
        ox, oy = self.offset
        lx, ly = x - ox, y - oy
        h, w = self.data.shape[:2]
        if 0 <= lx < w and 0 <= ly < h:
            self._cursor_mk.set_data([lx], [ly])
            self._cursor_mk.set_visible(True)
        else:
            self._cursor_mk.set_data([], [])
            self._cursor_mk.set_visible(False)
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)

    def clear_cursor(self):
        if self._cursor_mk is not None:
            self._cursor_mk.set_data([], [])
            self._cursor_mk.set_visible(False)
            _debounce(self, "_draw_timer", self.draw_idle, ms=40)

    # ---------- ROI/Line tools ----------
    def set_roi_enabled(self, enabled=True):
        if enabled and self.rect_selector is None:
            self.rect_selector = RectangleSelector(
                self.ax,
                self._on_rect_select,
                useblit=True,
                button=[1],
                minspanx=5,
                minspany=5,
                spancoords="pixels",
            )
        if self.rect_selector:
            self.rect_selector.set_active(enabled)

    def _on_rect_select(self, eclick, erelease):
        if eclick.xdata is None or erelease.xdata is None:
            return
        x1, y1 = int(eclick.xdata), int(eclick.ydata)
        x2, y2 = int(erelease.xdata), int(erelease.ydata)
        """Set highlight indices."""
        self.roi_selected.emit((x1, y1, x2, y2))

    def _connect_mouse(self):
        self.mcid_press = self.figure.canvas.mpl_connect("button_press_event", self._on_press)
        self.mcid_move = self.figure.canvas.mpl_connect("motion_notify_event", self._on_motion)
        self.mcid_rel = self.figure.canvas.mpl_connect("button_release_event", self._on_release)

    def _on_press(self, event):
        if event.inaxes != self.ax:
            return
        if event.button == 1:
            self.line_start = (event.xdata, event.ydata)
            self._ensure_tmp_line()

    def _ensure_tmp_line(self):
        if not hasattr(self, "_tmp_line"):
            (self._tmp_line,) = self.ax.plot([], [], linestyle="--")

    def _on_motion(self, event):
        if self.line_start is None or event.inaxes != self.ax:
            return
        self._ensure_tmp_line()
        x0, y0 = self.line_start
        self._tmp_line.set_data([x0, event.xdata], [y0, event.ydata])
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)

    def _on_release(self, event):
        if self.line_start is None or event.inaxes != self.ax:
            return
        x0, y0 = self.line_start
        x1, y1 = event.xdata, event.ydata
        self.line_start = None
        if hasattr(self, "_tmp_line"):
            self._tmp_line.remove()
            delattr(self, "_tmp_line")
        if x0 is not None and y0 is not None and x1 is not None and y1 is not None:
            self.line_selected.emit((int(x0), int(y0), int(x1), int(y1)))

    def set_highlight_indices(self, indices, full_shape=None):
        """Accept a set of FULL-image flat indices and reuse the
        existing set_highlight_points([(x,y), ...]) overlay."""
        if not indices:
            self.set_highlight_points([])
            return

        # Resolve full image shape
        H_full, W_full = (None, None)
        if full_shape is not None:
            H_full, W_full = full_shape
        else:
            # ask controller if available
            parent = self.parent() if callable(getattr(self, "parent", None)) else None
            if parent is not None and hasattr(parent, "maps_shape"):
                H_full, W_full = parent.maps_shape
        """Combine.

        """

        if not H_full or not W_full:
            self.set_highlight_points([])
            return

        # Convert flat -> (x,y)
        pts = []
        """Set data."""
        for i in indices:
            i = int(i)
            if 0 <= i < H_full * W_full:
                y, x = divmod(i, W_full)
                pts.append((x, y))

        self.set_highlight_points(pts)


# -----------------------------
# RGB map panel
# -----------------------------
class RGBMapView(ElementMapView):
    roi_selected = pyqtSignal(tuple)
    line_selected = pyqtSignal(tuple)

    def __init__(self, title="RGB Map"):
        super().__init__(title=title)
        self.data_triplet = None

    @staticmethod
    def _normalize(d):
        if d is None or d.size == 0:
            return d
        dmin = np.nanmin(d)
        dmax = np.nanmax(d)
        if not np.isfinite(dmin) or not np.isfinite(dmax) or dmax == dmin:
            return np.zeros_like(d, dtype=float)
        return (d - dmin) / (dmax - dmin)

    @staticmethod
    def combine(r, g, b, br=1.0, bg=1.0, bb=1.0):
        if r is None or g is None or b is None:
            return None
        if r.size == 0 or g.size == 0 or b.size == 0:
            return np.zeros((1, 1, 3), dtype=float)
        R = np.clip(RGBMapView._normalize(r) * br, 0, 1)
        G = np.clip(RGBMapView._normalize(g) * bg, 0, 1)
        B = np.clip(RGBMapView._normalize(b) * bb, 0, 1)
        return np.stack([R, G, B], axis=-1)

    def set_data(self, r, g, b, names=None, br=1.0, bg=1.0, bb=1.0, offset=(0, 0)):
        self.data_triplet = (r, g, b)
        self.offset = offset
        self.ax.clear()
        rgb = self.combine(r, g, b, br, bg, bb)
        """Set data."""
        self.im = self.ax.imshow(rgb)
        if names and len(names) >= 3:
            self.ax.set_title(f"{names[0]}-{names[1]}-{names[2]} (RGB) Map")
        else:
            self.ax.set_title(self.title)
        self.ax.set_xlim(-0.5, rgb.shape[1] - 0.5)
        self.ax.set_ylim(rgb.shape[0] - 0.5, -0.5)
        if getattr(self, "_last_highlight", None):
            self._ensure_hl_artist()
            self._update_hl_artist()
        self._ensure_cursor()
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)

    def _update_hl_artist(self):
        if self.data_triplet is None:
            return
        self._ensure_hl_artist()
        r, g, b = self.data_triplet
        h, w = r.shape[:2]
        ox, oy = self.offset
        xs, ys = [], []
        for x, y in self._last_highlight or []:
            lx, ly = x - ox, y - oy
            if 0 <= lx < w and 0 <= ly < h:
                xs.append(lx)
                ys.append(ly)
        if xs:
            self._hl_scatter.set_offsets(np.c_[xs, ys])
            self._hl_scatter.set_visible(self._hl_visible)
        else:
            self._hl_scatter.set_offsets(np.empty((0, 2)))
            self._hl_scatter.set_visible(False)

    def refresh_brightness(self, br=1.0, bg=1.0, bb=1.0, names=None):
        if self.data_triplet is None:
            return
        r, g, b = self.data_triplet
        self.set_data(r, g, b, names, br, bg, bb)


# -----------------------------
# Gray image panel (SEM or similar)
# -----------------------------
class GrayImageView(ElementMapView):
    def __init__(self, title="Image"):
        super().__init__(title=title)

    def set_data(self, img, title=None, offset=(0, 0)):
        self.data = img
        self.offset = offset
        self.ax.clear()
        self.im = self.ax.imshow(img, cmap="gray")
        self.ax.set_title(title or self.title)
        self.ax.set_xlim(-0.5, img.shape[1] - 0.5)
        self.ax.set_ylim(img.shape[0] - 0.5, -0.5)
        if getattr(self, "_last_highlight", None):
            self._ensure_hl_artist()
            self._update_hl_artist()
        self._ensure_cursor()
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)

    # Re-added to fix AttributeError when AnalysisStep calls set_zoom
    def set_zoom(self, x1, y1, x2, y2, base_img):
        crop = base_img[y1:y2, x1:x2]
        self.set_data(crop, title=self.title, offset=(x1, y1))


# -----------------------------
# Simple Ternary Plot Dialog
# -----------------------------


# Matplotlib imports are placed inside methods where possible to avoid
# early backend binding


class TernaryDialog(QDialog):
    # Emits list of FULL-image (x,y) pixels selected by lasso
    lasso_pixels_xy = pyqtSignal(list)

    def __init__(self, parent=None, on_select_indices=None):
        super().__init__(parent)
        self._owner = parent
        self.on_select_indices = on_select_indices

        # Canvas / axes
        self.fig = None
        self.canvas = None
        self.ax = None

        # Data backing the scatter currently shown
        self._xy = None  # (N,2) plotted coords in ternary axes
        # (N,) indices mapping to *full-image* flat indices (preferred)
        self._flat_indices = None
        self._flat_indices_full = None  # if you prefer to keep explicit FULL-flat here
        # (H_roi, W_roi) ROI shape (for legacy conversion)
        self._roi_shape = None
        # (x0,y0) top-left of ROI in FULL image (legacy)
        self._roi_origin = (0, 0)
        self._triangle_path = None

        # Selection & overlays
        self._lasso = None
        self._labels = ("A", "B", "C")

        self._sel_mask = None  # boolean mask over self._xy
        self._sel_scatter = None  # overlay for selected points (ternary axes)
        """Set labels."""

        # “ElementMapView”-style highlight overlay (optional second overlay)
        self._hl_scatter = None
        self._hl_visible = True
        self._last_highlight_xy = []

        # Build layout and canvas lazily
        self._ensure_canvas_axes()
        lay = self.layout() or QVBoxLayout(self)
        lay.setContentsMargins(6, 6, 6, 6)
        if self.canvas.parent() is None:
            lay.addWidget(self.canvas)

        self.setWindowTitle("Triangular Plot")

    # ------------- Canvas / Axes -------------
    def _ensure_canvas_axes(self):
        from matplotlib.figure import Figure

        """Update points.

        """
        try:
            from matplotlib.backends.backend_qt5agg import (
                FigureCanvasQTAgg as FigureCanvas,
            )
        except Exception:
            from matplotlib.backends.backend_qtagg import (
                FigureCanvasQTAgg as FigureCanvas,
            )

        if self.fig is None:
            self.fig = Figure(figsize=(4, 4), tight_layout=True)
        if self.canvas is None:
            self.canvas = FigureCanvas(self.fig)
        if self.ax is None:
            self.ax = self.fig.add_subplot(111)

    def _setup_axes(self):
        self._ensure_canvas_axes()
        self.ax.clear()

        # Triangle vertices in 2D cartesian
        A = np.array([0.5, math.sqrt(3) / 2.0])  # top
        B = np.array([0.0, 0.0])  # left
        C = np.array([1.0, 0.0])  # right
        tri = np.vstack([A, B, C, A])

        self.ax.plot(tri[:, 0], tri[:, 1], linewidth=1.0)
        self.ax.set_xticks([])
        self.ax.set_yticks([])
        self.ax.set_xlim(-0.05, 1.05)
        self.ax.set_ylim(-0.05, math.sqrt(3) / 2.0 + 0.05)

        self.ax.text(
            A[0],
            A[1] + 0.04,
            self._labels[0],
            ha="center",
            va="bottom",
            fontsize=10,
        )
        self.ax.text(
            B[0] - 0.02,
            B[1] - 0.02,
            self._labels[1],
            ha="right",
            va="top",
            fontsize=10,
        )
        self.ax.text(
            C[0] + 0.02,
            C[1] - 0.02,
            self._labels[2],
            ha="left",
            va="top",
            fontsize=10,
        )
        self.ax.set_title("Triangular Plot")

        # boundary path for containment checks
        self._triangle_path = Path(tri[:-1, :])
        self.canvas.draw_idle()

    def set_labels(self, labels3):
        if labels3 and len(labels3) == 3:
            self._labels = tuple(labels3)
        self._setup_axes()

    # ------------- Data prep -------------
    @staticmethod
    def _to_cartesian(a, b, c):
        """Map (a,b,c) with a+b+c=1 to 2D point in triangle space."""
        A = np.array([0.5, math.sqrt(3) / 2.0])
        B = np.array([0.0, 0.0])
        C = np.array([1.0, 0.0])
        return a * A + b * B + c * C

    def _ensure_selection_mask(self):
        if self._xy is None:
            self._sel_mask = None
            return
        n = len(self._xy)
        if self._sel_mask is None or self._sel_mask.shape != (n,):
            self._sel_mask = np.zeros((n,), dtype=bool)

    def update_points(
        self,
        a,
        b,
        c,
        roi_shape=None,
        roi_origin=(0, 0),
        flat_indices_full=None,
    ):
        self._ensure_canvas_axes()
        self._setup_axes()

        a = np.asarray(a, dtype=float)
        b = np.asarray(b, dtype=float)
        c = np.asarray(c, dtype=float)

        n = a.size
        if roi_shape is None:
            side = int(np.sqrt(n) + 0.5)
            roi_shape = (side, side)

        # Shared finite mask
        finite_mask = np.isfinite(a) & np.isfinite(b) & np.isfinite(c)
        if not np.any(finite_mask):
            self._xy = None
            self._flat_indices = None
            self._flat_indices_full = None
            self._roi_shape = roi_shape
            self._roi_origin = roi_origin
            self.canvas.draw_idle()
            return

        def minmax(x):
            xmin = np.nanmin(x)
            xmax = np.nanmax(x)
            if not np.isfinite(xmin) or not np.isfinite(xmax) or xmax == xmin:
                return np.zeros_like(x, dtype=float)
            return (x - xmin) / (xmax - xmin)

        af = minmax(a[finite_mask])
        bf = minmax(b[finite_mask])
        cf = minmax(c[finite_mask])
        s = af + bf + cf
        pos_mask = s > 0
        if not np.any(pos_mask):
            self._xy = None
            self._flat_indices = None
            self._flat_indices_full = None
            self._roi_shape = roi_shape
            self._roi_origin = roi_origin
            self.canvas.draw_idle()
            return

        af = af[pos_mask] / s[pos_mask]
        bf = bf[pos_mask] / s[pos_mask]
        cf = cf[pos_mask] / s[pos_mask]

        # Index maps
        idx_all = np.arange(n)
        idx_kept = idx_all[finite_mask][pos_mask]
        self._flat_indices = idx_kept.copy()  # by default store "ROI-flat" (0..n-1 of this view)
        """Set highlight visible."""
        self._flat_indices_full = None
        if flat_indices_full is not None:
            flat_indices_full = np.asarray(flat_indices_full, dtype=np.int64)
            self._flat_indices_full = flat_indices_full[finite_mask][pos_mask]

        # Cartesian coords in ternary
        xy = np.array([self._to_cartesian(ai, bi, ci) for ai, bi, ci in zip(af, bf, cf, strict=False)])

        # Keep only those inside triangle
        if self._triangle_path is not None:
            inside = self._triangle_path.contains_points(xy)
            xy = xy[inside]
            self._flat_indices = self._flat_indices[inside]
            if self._flat_indices_full is not None:
                self._flat_indices_full = self._flat_indices_full[inside]

        # Draw base points
        self.ax.scatter(xy[:, 0], xy[:, 1], s=6, alpha=0.6, zorder=3)
        self.canvas.draw_idle()

        # Save arrays and ROI info
        self._xy = xy
        self._roi_shape = roi_shape
        self._roi_origin = roi_origin

        # Reset selection mask and overlays
        """Highlight from fullimg points.

        """
        self._sel_mask = None
        self._ensure_sel_artist()
        self._ensure_hl_artist()
        self.set_highlight_points([])

    # ------------- Overlays -------------
    def _ensure_sel_artist(self):
        if self._sel_scatter is None or getattr(self._sel_scatter, "axes", None) is None:
            self._sel_scatter = self.ax.scatter(
                [],
                [],
                s=12,
                facecolors="none",
                edgecolors="k",
                linewidths=1.0,
                zorder=6,
            )

    def _update_local_selection_overlay(self):
        self._ensure_selection_mask()
        self._ensure_sel_artist()
        if self._sel_mask is None or self._xy is None:
            self._sel_scatter.set_visible(False)
            self.canvas.draw_idle()
            return
        pts = self._xy[self._sel_mask]
        if pts.size == 0:
            self._sel_scatter.set_offsets(np.empty((0, 2)))
            self._sel_scatter.set_visible(False)
        else:
            self._sel_scatter.set_offsets(pts)
            self._sel_scatter.set_visible(True)
        self.canvas.draw_idle()

    # ElementMapView-style “external highlight” overlay (optional second layer)
    def _ensure_hl_artist(self):
        if self._hl_scatter is None or getattr(self._hl_scatter, "axes", None) is None:
            self._hl_scatter = self.ax.scatter(
                [],
                [],
                s=18,
                facecolors="none",
                edgecolors="yellow",
                linewidths=1.0,
                zorder=7,
            )
            self._hl_visible = True
            self._last_highlight_xy = []

    def set_highlight_visible(self, flag: bool):
        self._hl_visible = bool(flag)
        if self._hl_scatter is not None:
            self._hl_scatter.set_visible(self._hl_visible and len(self._last_highlight_xy) > 0)
            self.canvas.draw_idle()

    def _update_hl_artist(self):
        self._ensure_hl_artist()
        pts = self._last_highlight_xy or []
        if len(pts) > 0:
            self._hl_scatter.set_offsets(np.asarray(pts))
            self._hl_scatter.set_visible(self._hl_visible)
        else:
            self._hl_scatter.set_offsets(np.empty((0, 2)))
            self._hl_scatter.set_visible(False)
        self.canvas.draw_idle()

    def set_highlight_points(self, coords_xy_list):
        """Directly set highlight points in ternary axes (for external
        sync)."""
        self._last_highlight_xy = coords_xy_list or []
        self._update_hl_artist()

    def highlight_from_fullimg_indices(self, idxs_full: set, parent=None):
        """Overlay (externally) from a set of FULL-image flat
        indices."""
        if self._xy is None or (self._flat_indices is None and self._flat_indices_full is None):
            self.set_highlight_points([])
            return
        flat = self._flat_indices_full if self._flat_indices_full is not None else self._flat_indices
        if not idxs_full or flat is None:
            self.set_highlight_points([])
            return
        mask = np.isin(flat, np.fromiter(idxs_full, dtype=np.int64))
        pts = self._xy[mask] if np.any(mask) else np.empty((0, 2))
        self.set_highlight_points(pts)

    def highlight_from_fullimg_points(self, full_coords, parent=None):
        """Overlay (externally) from FULL-image (x,y) pixels."""
        if not full_coords:
            self.set_highlight_points([])
            return
        if parent is None:
            parent = self._owner
        H_full, W_full = getattr(parent, "maps_shape", (None, None)) if parent is not None else (None, None)
        if not H_full or not W_full:
            self.set_highlight_points([])
            return
        idxs_full = {
            int(y) * int(W_full) + int(x) for (x, y) in full_coords if 0 <= int(x) < W_full and 0 <= int(y) < H_full
        }
        self.highlight_from_fullimg_indices(idxs_full, parent=parent)

    # ------------- Lasso / Interaction -------------
    def enable_lasso(self, enabled: bool):
        self._ensure_canvas_axes()
        if enabled and self._lasso is None:
            self._lasso = _make_lasso_selector(self.ax, self._on_lasso)
        elif not enabled and self._lasso is not None:
            self._lasso.disconnect_events()
            self._lasso = None

    def _on_lasso(self, verts):
        """Lasso behavior:
          - no modifier: replace selection
          - Shift held:  add to selection
          - Ctrl held:   subtract from selection
        Also updates parent.selected_pixels and parent.selection.set(idxs).
        """
        if self._xy is None or self._flat_indices is None:
            return

        mods = QApplication.keyboardModifiers()
        add_mode = bool(mods & Qt.ShiftModifier)
        sub_mode = bool(mods & Qt.ControlModifier)

        lasso_poly = Path(verts)
        in_lasso = lasso_poly.contains_points(self._xy)
        if self._triangle_path is not None:
            in_lasso &= self._triangle_path.contains_points(self._xy)

        if not np.any(in_lasso):
            # No change if empty lasso (like correlation behavior)
            return

        # Update local mask (correlation-like UX)
        self._ensure_selection_mask()
        if sub_mode:
            self._sel_mask[in_lasso] = False
        elif add_mode:
            self._sel_mask[in_lasso] = True
        else:
            self._sel_mask[:] = False
            self._sel_mask[in_lasso] = True

        # Local overlay
        self._update_local_selection_overlay()

        # Convert selected ternary rows -> FULL (x,y) and FULL flat indices
        # Prefer explicit full-flat if available
        if self._flat_indices_full is not None:
            sel_full_flat = self._flat_indices_full[self._sel_mask]
            full_coords = None  # not needed if parent only consumes indices
        else:
            # Legacy: convert via ROI shape + origin
            sel_roi_flat = self._flat_indices[self._sel_mask]
            if self._roi_shape is None:
                return
            H_roi, W_roi = self._roi_shape
            rows, cols = np.unravel_index(sel_roi_flat, (H_roi, W_roi))
            x0, y0 = self._roi_origin
            full_coords = [(int(x0 + int(c)), int(y0 + int(r))) for r, c in zip(rows, cols, strict=False)]
            # need full-flat for downstream
            owner = self._owner
            H_full, W_full = getattr(owner, "maps_shape", (None, None)) if owner else (None, None)
            if H_full and W_full:
                sel_full_flat = np.array(
                    [int(y) * int(W_full) + int(x) for (x, y) in full_coords],
                    dtype=np.int64,
                )
            else:
                sel_full_flat = np.array([], dtype=np.int64)

        # Broadcast to parent
        parent = self._owner
        if parent is not None:
            # Keep pixel list for point-data workflows (if we have coords)
            try:
                if full_coords is None and sel_full_flat.size > 0:
                    # derive coords from full-flat if maps_shape exists
                    Hf, Wf = getattr(parent, "maps_shape", (None, None))
                    if Hf and Wf:
                        ys, xs = np.divmod(sel_full_flat, Wf)
                        full_coords = list(zip(xs.tolist(), ys.tolist(), strict=False))
                parent.selected_pixels = full_coords or []
            except Exception:
                parent.selected_pixels = full_coords or []

            # Unified selection manager for maps + correlation
            if hasattr(parent, "selection") and sel_full_flat.size > 0:
                try:
                    parent.selection.set(set(sel_full_flat.tolist()))
                except Exception:
                    pass
            else:
                # Fallback: manual refresh if selection manager not used
                if hasattr(parent, "_apply_highlights_everywhere"):
                    parent._apply_highlights_everywhere()
                sc = getattr(parent, "scatter", None)
                if sc is not None and hasattr(sc, "set_highlight_indices"):
                    sc.set_highlight_indices(set(sel_full_flat.tolist()))

        # Emit signal for any external listeners
        try:
            # If we have coords, emit them; else derive quickly for listeners

            emit_coords = full_coords
            """Ranks."""
            if emit_coords is None and sel_full_flat.size > 0:
                Hf, Wf = getattr(self._owner, "maps_shape", (None, None)) if self._owner else (None, None)
                if Hf and Wf:
                    ys, xs = np.divmod(sel_full_flat, Wf)
                    emit_coords = list(zip(xs.tolist(), ys.tolist(), strict=False))
                else:
                    emit_coords = []
            self.lasso_pixels_xy.emit(emit_coords or [])
        except Exception:
            pass


# -----------------------------
# Scatter (correlation) panel
# -----------------------------
class ScatterView(BasePlotWidget):
    # Emits ROI-local pixel coords (x,y) in the currently displayed ROI
    selection_made = pyqtSignal(list)

    def __init__(self, title="Correlation Plot"):
        super().__init__()
        self.title = title
        self.scatter_limits = None
        self._x_flat = None
        self._y_flat = None
        self._shape = None
        self._sel_scatter = None
        self._lasso = None
        self._stats_box = None
        self._sx = None
        self._sy = None
        self._mc_enabled = False
        self._mc_iters = 1000
        self._mc_cache = None
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    @staticmethod
    def _pearson_with_ci(x, y):
        """Pearson r and 95% CI via Fisher z-transform.

        Returns (r, lo, hi, n).
        """
        x = np.asarray(x)
        y = np.asarray(y)
        mask = np.isfinite(x) & np.isfinite(y)
        x = x[mask]
        y = y[mask]
        n = x.size
        if n < 4:
            return (np.nan, np.nan, np.nan, n)
        # center
        x0 = x - x.mean()
        """Set data."""
        y0 = y - y.mean()
        denom = np.sqrt(np.sum(x0 * x0) * np.sum(y0 * y0))
        r = np.sum(x0 * y0) / denom if denom != 0 else np.nan
        if not np.isfinite(r):
            return (np.nan, np.nan, np.nan, n)
        # CI
        z = np.arctanh(np.clip(r, -0.999999, 0.999999))
        se = 1.0 / np.sqrt(n - 3)
        z_lo = z - 1.96 * se
        z_hi = z + 1.96 * se
        return (r, np.tanh(z_lo), np.tanh(z_hi), n)

    @staticmethod
    def _spearman_rho(x, y):
        """Spearman correlation (Pearson of ranks), no ties correction
        beyond average ranks."""
        x = np.asarray(x)
        y = np.asarray(y)
        mask = np.isfinite(x) & np.isfinite(y)
        x = x[mask]
        y = y[mask]
        n = x.size
        if n < 2:
            return np.nan

        # average ranks
        def ranks(a):
            order = np.argsort(a)
            r = np.empty_like(order, dtype=float)
            r[order] = np.arange(len(a), dtype=float)
            # average ranks for ties:
            # find equal runs in sorted a
            sa = a[order]
            i = 0
            while i < len(sa):
                j = i + 1
                while j < len(sa) and sa[j] == sa[i]:
                    j += 1
                # average the ranks [i, j)
                avg = (i + j - 1) / 2.0
                r[order[i:j]] = avg
                i = j
            return r

        rx = ranks(x)
        ry = ranks(y)
        rx -= rx.mean()
        ry -= ry.mean()
        denom = np.sqrt(np.sum(rx * rx) * np.sum(ry * ry))
        return (np.sum(rx * ry) / denom) if denom != 0 else np.nan

    def set_data(self, x_arr, y_arr, xlabel="X", ylabel="Y", remember_limits=False):
        """Load data into the scatter view, build the finite mask, store
        flattened vectors, sync per-pixel uncertainties (σ) to the same
        mask, and redraw."""
        # Ensure attrs exist
        if not hasattr(self, "_scat"):
            self._scat = None
        if not hasattr(self, "_stats_box"):
            self._stats_box = None
        if not hasattr(self, "_flat_indices"):
            self._flat_indices = None
        if not hasattr(self, "_sx_full"):
            self._sx_full = None
        if not hasattr(self, "_sy_full"):
            self._sy_full = None
        if not hasattr(self, "_sx_flat"):
            self._sx_flat = None
        if not hasattr(self, "_sy_flat"):
            self._sy_flat = None
        if not hasattr(self, "_mc_enabled"):
            self._mc_enabled = False
        """Enable lasso.

        """
        if not hasattr(self, "_mc_iters"):
            self._mc_iters = 1000
        if not hasattr(self, "_mc_cache"):
            self._mc_cache = None

        # Keep previous view limits if requested
        prev_xlim = prev_ylim = None
        if remember_limits and hasattr(self.ax, "get_xlim"):
            prev_xlim = self.ax.get_xlim()
            prev_ylim = self.ax.get_ylim()

        # Convert to float arrays
        x2 = np.asarray(x_arr, dtype=float)
        y2 = np.asarray(y_arr, dtype=float)

        self._shape = x2.shape[:2]
        # Finite mask (your scatter should reflect only valid points)
        mask2d = np.isfinite(x2) & np.isfinite(y2)
        # Store mapping from full ravel to current valid subset
        self._flat_indices = np.nonzero(mask2d.ravel())[0]

        # Flattened, masked data actually plotted
        self._x_flat = x2.ravel()[self._flat_indices]
        self._y_flat = y2.ravel()[self._flat_indices]

        # Sync σ to the same mask (if already provided)
        self._sx_flat = None
        self._sy_flat = None
        try:
            if (self._sx_full is not None) and (self._sy_full is not None):
                self._sx_flat = np.asarray(self._sx_full, dtype=float).ravel()[self._flat_indices]
                self._sy_flat = np.asarray(self._sy_full, dtype=float).ravel()[self._flat_indices]
        except Exception:
            # If shapes don't match, drop σ to keep UI robust
            self._sx_flat = None

            self._sy_flat = None

        # Clear MC cache whenever data changes
        self._mc_cache = None

        # Create/Update the artist
        if self._scat is None:
            self._scat = self.ax.scatter(
                self._x_flat,
                self._y_flat,
                s=5,
                alpha=0.7,
                edgecolors="none",
                rasterized=True,
                zorder=3,
            )
        else:
            # set_offsets expects Nx2
            if self._x_flat.size:
                self._scat.set_offsets(np.c_[self._x_flat, self._y_flat])
            else:
                self._scat.set_offsets(np.empty((0, 2)))

        # Labels & grids
        self.ax.set_xlabel(str(xlabel))
        self.ax.set_ylabel(str(ylabel))
        try:
            self.ax.grid(True, which="major", alpha=0.2)
        except Exception:
            pass

        # Autoscale if not keeping previous limits
        if (prev_xlim is None) or (prev_ylim is None):
            if self._x_flat.size > 0:
                # Pad 2%
                xmin, xmax = float(np.min(self._x_flat)), float(np.max(self._x_flat))
                ymin, ymax = float(np.min(self._y_flat)), float(np.max(self._y_flat))
                if np.isfinite([xmin, xmax, ymin, ymax]).all():
                    dx = max((xmax - xmin) * 0.02, 1e-12)
                    dy = max((ymax - ymin) * 0.02, 1e-12)
                    self.ax.set_xlim(xmin - dx, xmax + dx)
                    self.ax.set_ylim(ymin - dy, ymax + dy)
        else:
            # Restore previous limits
            self.ax.set_xlim(prev_xlim)
            self.ax.set_ylim(prev_ylim)

        # Draw stats box
        self._render_stats_box()
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)

    def enable_lasso(self, enabled=True):
        if enabled:
            if self._lasso is None:
                self._lasso = _make_lasso_selector(self.ax, self._on_lasso_select)
            else:
                self._lasso.set_active(True)
        else:
            if self._lasso is not None:

                self._lasso.set_active(False)

    def _on_lasso_select(self, verts):
        if self._x_flat is None or self._y_flat is None or self._shape is None:
            return
        path = Path(verts)
        pts = np.c_[self._x_flat, self._y_flat]
        mask = path.contains_points(pts)
        idxs = np.nonzero(mask)[0]  # indices in the *subset* (x_flat/y_flat)

        # ensure selection overlay exists
        if self._sel_scatter is None or getattr(self._sel_scatter, "axes", None) is None:
            self._sel_scatter = self.ax.scatter(
                [],
                [],
                s=12,
                facecolors="none",
                edgecolors="k",
                linewidths=1.0,
                zorder=6,
            )

        if idxs.size:
            self._sel_scatter.set_offsets(np.c_[self._x_flat[idxs], self._y_flat[idxs]])
            self._sel_scatter.set_visible(True)
        else:
            self._sel_scatter.set_offsets(np.empty((0, 2)))
            self._sel_scatter.set_visible(False)
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)

        # NEW: map subset indices -> FULL flat indices -> (x,y) in FULL image
        full_idxs = self._flat_indices[idxs]  # FULL-image flat indices
        rows, cols = np.unravel_index(full_idxs, self._shape)  # row=y, col=x
        coords = [(int(c), int(r)) for r, c in zip(rows, cols, strict=False)]
        self.selection_made.emit(coords)

    def set_highlight_indices(self, indices):
        """Accept a set of FULL-image flat indices and draw them
        highlighted on the current scatter (which shows only a subset).

        We map FULL → subset positions.
        """
        if self._x_flat is None or self._y_flat is None or self._flat_indices is None:
            return
        if self._sel_scatter is None or getattr(self._sel_scatter, "axes", None) is None:
            self._sel_scatter = self.ax.scatter(
                [],
                [],
                s=12,
                facecolors="none",
                edgecolors="k",
                linewidths=1.0,
                zorder=6,
            )
        if not indices:
            self._sel_scatter.set_offsets(np.empty((0, 2)))
            self._sel_scatter.set_visible(False)
            _debounce(self, "_draw_timer", self.draw_idle, ms=40)
            return

        # NEW: convert provided FULL-image flat indices → positions in current
        # subset
        want = np.fromiter((int(i) for i in indices), dtype=np.int64)
        if want.size == 0:
            self._sel_scatter.set_offsets(np.empty((0, 2)))
            self._sel_scatter.set_visible(False)
            _debounce(self, "_draw_timer", self.draw_idle, ms=40)
            return

        mask_in_subset = np.isin(self._flat_indices, want)
        idxs_subset = np.nonzero(mask_in_subset)[0]

        if idxs_subset.size == 0:
            self._sel_scatter.set_offsets(np.empty((0, 2)))
            self._sel_scatter.set_visible(False)
            _debounce(self, "_draw_timer", self.draw_idle, ms=40)
            return

        self._sel_scatter.set_offsets(np.c_[self._x_flat[idxs_subset], self._y_flat[idxs_subset]])
        self._sel_scatter.set_visible(True)
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)

    def set_uncertainty(self, sx, sy):
        self._sx_full = None if sx is None else np.asarray(sx, dtype=float)
        self._sy_full = None if sy is None else np.asarray(sy, dtype=float)
        self._sx_flat = None
        self._sy_flat = None
        self._mc_cache = None
        self._sync_sigma_flat()

    def _sync_sigma_flat(self):
        if (self._flat_indices is None) or (self._sx_full is None) or (self._sy_full is None):
            return
        # Accept either 2D (H,W) or already-flat matching the full image size
        sx_flat_all = self._sx_full.ravel()
        sy_flat_all = self._sy_full.ravel()

        # Guard: need same base-size as original maps
        try:
            self._sx_flat = sx_flat_all[self._flat_indices]
            self._sy_flat = sy_flat_all[self._flat_indices]
        except Exception:
            # If shapes still mismatch, drop σ to avoid crashes
            self._sx_flat = None
            self._sy_flat = None

    def enable_mc(self, enabled: bool, iters: int = None):
        """Enable/disable Monte Carlo propagation, optionally update
        iterations."""
        self._mc_enabled = bool(enabled)
        if iters is not None:
            try:
                self._mc_iters = max(100, int(iters))
            except Exception:
                pass
        # Clear old cache so it recomputes next render
        self._mc_cache = None
        # Immediately refresh stats
        self._render_stats_box()
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)

    def _pearson(self, x, y):
        mask = np.isfinite(x) & np.isfinite(y)
        x = x[mask]
        y = y[mask]
        if x.size < 2:
            return np.nan
        x0 = x - x.mean()
        y0 = y - y.mean()
        denom = np.sqrt((x0 * x0).sum() * (y0 * y0).sum())
        return (x0 * y0).sum() / denom if denom > 0 else np.nan

    def _mc_pearson(
        self,
        x,
        y,
        sx,
        sy,
        iters=1000,
        seed=123,
        chunk=250,
        stop_eps=0.01,
        min_iters=400,
    ):
        """Chunked MC Pearson; stops early when CI half-width <
        stop_eps.

        Memory-light: holds only one (N, chunk) at a time.
        """
        rng = np.random.default_rng(seed)
        x = np.asarray(x, float)
        y = np.asarray(y, float)
        sx = np.asarray(sx, float)
        sy = np.asarray(sy, float)
        m = np.isfinite(x) & np.isfinite(y) & np.isfinite(sx) & np.isfinite(sy)
        x = x[m]
        y = y[m]
        sx = sx[m]
        sy = sy[m]
        if x.size < 8:
            return None

        rs = []
        done = 0
        while done < iters:
            k = min(chunk, iters - done)
            # generate k realizations
            X = rng.normal(x[:, None], sx[:, None], size=(x.size, k))
            Y = rng.normal(y[:, None], sy[:, None], size=(y.size, k))
            # center each realization (column-wise)
            X -= X.mean(axis=0, keepdims=True)
            Y -= Y.mean(axis=0, keepdims=True)
            num = (X * Y).sum(axis=0)
            den = np.sqrt((X * X).sum(axis=0) * (Y * Y).sum(axis=0))
            r = num / np.maximum(den, 1e-20)
            rs.append(r)
            done += k

            # early stop after min_iters when CI tight enough
            if done >= min_iters:
                rr = np.concatenate(rs)
                lo, hi = np.percentile(rr[np.isfinite(rr)], [2.5, 97.5])
                if (hi - lo) * 0.5 < stop_eps:
                    break

        rr = np.concatenate(rs)
        rr = rr[np.isfinite(rr)]
        if rr.size == 0:
            return None
        lo, hi = np.percentile(rr, [2.5, 97.5])
        return float(rr.mean()), float(lo), float(hi)

    # def _render_stats_box(self):
    #     if self._x_flat is None or self._y_flat is None:
    #         return
    #     r, lo, hi, n = self._pearson_with_ci(self._x_flat, self._y_flat)
    #     rho = self._spearman_rho(self._x_flat, self._y_flat)

    #     txt = []
    #     txt.append(f"n = {n}")
    #     if np.isfinite(r):
    #         txt.append(f"Pearson r = {r:.3f}  (95% CI: {lo:.3f}–{hi:.3f})")
    #     else:
    #         txt.append("Pearson r = NA")
    #     if np.isfinite(rho):
    #         txt.append(f"Spearman ρ = {rho:.3f}")
    #     else:
    #         txt.append("Spearman ρ = NA")

    #     text = "\n".join(txt)
    #     if (
    #         self._stats_box is None
    #         or getattr(self._stats_box, "axes", None) is None
    #     ):
    #         self._stats_box = self.ax.text(
    #             0.02,
    #             0.98,
    #             text,
    #             transform=self.ax.transAxes,
    #             ha="left",
    #             va="top",
    #             fontsize=9,
    #             bbox=dict(
    #                 boxstyle="round",
    #                 facecolor="white",
    #                 alpha=0.6,
    #                 linewidth=0.5,
    #             ),
    #             zorder=8,
    #         )
    #     else:
    #         self._stats_box.set_text(text)

    def _render_stats_box(self):
        if not hasattr(self, "_x_flat") or self._x_flat is None or self._x_flat.size < 2:
            text = "n = 0"
            if self._stats_box is None or getattr(self._stats_box, "axes", None) is None:
                self._stats_box = self.ax.text(
                    0.02,
                    0.98,
                    text,
                    transform=self.ax.transAxes,
                    ha="left",
                    va="top",
                    fontsize=9,
                    bbox=dict(
                        boxstyle="round",
                        facecolor="white",
                        alpha=0.6,
                        linewidth=0.5,
                    ),
                    zorder=8,
                )
            else:
                self._stats_box.set_text(text)
            return

        x = self._x_flat
        y = self._y_flat
        n = int(min(x.size, y.size))

        # --- Pearson r (classical) ---
        def _pearson_xy(a, b):
            am = a - np.nanmean(a)
            bm = b - np.nanmean(b)
            denom = np.sqrt(np.nansum(am * am) * np.nansum(bm * bm))
            return np.nan if denom <= 0 else float(np.nansum(am * bm) / denom)

        r = _pearson_xy(x, y)

        # Fisher 95% CI for r (classical)
        def _fisher_ci(r_val, n_obs, zcrit=1.96):
            if not np.isfinite(r_val) or n_obs < 4:
                return (np.nan, np.nan)
            r_clip = np.clip(r_val, -0.999999, 0.999999)
            z = np.arctanh(r_clip)
            se = 1.0 / np.sqrt(max(n_obs - 3, 1))
            z_lo, z_hi = z - zcrit * se, z + zcrit * se
            return (float(np.tanh(z_lo)), float(np.tanh(z_hi)))

        r_lo, r_hi = _fisher_ci(r, n)

        # --- Spearman ρ (rank correlation) ---
        def _spearman_rho(a, b):
            # rank using numpy (average rank for ties)
            def _ranks(v):
                v = np.asarray(v)
                order = np.argsort(v)
                ranks = np.empty_like(order, dtype=float)
                ranks[order] = np.arange(1, v.size + 1, dtype=float)
                # average ties
                # (simple tie-handling: average ranks of equal values)
                unique, inv, counts = np.unique(v, return_inverse=True, return_counts=True)
                if np.any(counts > 1):
                    cum = np.cumsum(counts)
                    starts = cum - counts + 1
                    means = (starts + cum) / 2.0
                    ranks = means[inv]
                return ranks

            rx = _ranks(a)
            ry = _ranks(b)
            return _pearson_xy(rx, ry)

        rho = _spearman_rho(x, y)

        # --- Weighted Pearson r_w (uses σ if available) ---
        rw = np.nan
        if (getattr(self, "_sx_flat", None) is not None) and (getattr(self, "_sy_flat", None) is not None):
            try:
                # prefer external helper if present
                if "_weighted_pearson" in globals():
                    w = 1.0 / np.maximum(self._sx_flat**2 + self._sy_flat**2, 1e-12)

                    rw = _weighted_pearson(x, y, w)
                else:
                    # inline weighted Pearson (fallback)
                    w = 1.0 / np.maximum(self._sx_flat**2 + self._sy_flat**2, 1e-12)
                    m = np.isfinite(x) & np.isfinite(y) & np.isfinite(w) & (w > 0)
                    if np.any(m):
                        xx, yy, ww = x[m], y[m], w[m]
                        W = np.sum(ww)
                        mx = np.sum(ww * xx) / W
                        my = np.sum(ww * yy) / W
                        cov = np.sum(ww * (xx - mx) * (yy - my)) / W
                        vx = np.sum(ww * (xx - mx) ** 2) / W
                        vy = np.sum(ww * (yy - my) ** 2) / W
                        denom = np.sqrt(vx * vy)
                        rw = float(cov / denom) if denom > 0 else np.nan
            except Exception:
                rw = np.nan

        # --- Monte-Carlo Pearson r (if enabled & σ available) ---
        mc_line = None
        if (
            bool(getattr(self, "_mc_enabled", False))
            and (getattr(self, "_sx_flat", None) is not None)
            and (getattr(self, "_sy_flat", None) is not None)
            and hasattr(self, "_mc_pearson")
        ):
            if self._mc_cache is None:
                try:
                    self._mc_cache = self._mc_pearson(
                        x,
                        y,
                        self._sx_flat,
                        self._sy_flat,
                        iters=int(getattr(self, "_mc_iters", 1000)),
                        seed=123,
                    )
                except Exception:
                    self._mc_cache = None
            if self._mc_cache is not None:
                rm, rlo, rhi = self._mc_cache
                mc_line = (
                    f"MC Pearson r: {rm:.3f} "
                    f"[2.5–97.5%: {rlo:.3f}–{rhi:.3f}] "
                    f"(N={int(getattr(self, '_mc_iters', 1000))})"
                )
            else:
                mc_line = "MC Pearson r: NA (insufficient σ)"

        # --- Compose stats text ----------------------------------------------
        lines = [f"n = {n}"]
        lines.append(f"Pearson r = {r:.3f}" + (f"  (95% CI: {r_lo:.3f}–{r_hi:.3f})" if np.isfinite(r_lo) else ""))
        lines.append(f"Spearman ρ = {rho:.3f}" if np.isfinite(rho) else "Spearman ρ = NA")
        if np.isfinite(rw):
            lines.append(f"Weighted Pearson r_w = {rw:.3f}")
        if mc_line:
            lines.append(mc_line)

        text = "\n".join(lines)

        # Create/update stats box
        if self._stats_box is None or getattr(self._stats_box, "axes", None) is None:
            self._stats_box = self.ax.text(
                0.02,
                0.98,
                text,
                transform=self.ax.transAxes,
                ha="left",
                va="top",
                fontsize=9,
                bbox=dict(
                    boxstyle="round",
                    facecolor="white",
                    alpha=0.6,
                    linewidth=0.5,
                ),
                zorder=8,
            )
        else:
            self._stats_box.set_text(text)

        # Redraw
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)


# -----------------------------
# Line scan panel
# -----------------------------
class LineScanView(BasePlotWidget):
    hover_index_changed = pyqtSignal(int)  # emits nearest index under cursor

    def __init__(self, title="Line Scan"):
        super().__init__()
        self.title = title
        self._xs = None
        self._ys_list = None
        self._lines = []
        self._vline = None
        self._markers = []
        self.mcid_move = self.figure.canvas.mpl_connect("motion_notify_event", self._on_motion)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

    def set_profiles(self, xs, profiles, labels):
        self.ax.clear()
        self._lines = []
        self._markers = []
        self._xs = np.asarray(xs)
        self._ys_list = [np.asarray(y) for y in profiles]
        for y, lab in zip(self._ys_list, labels, strict=False):
            (ln,) = self.ax.plot(self._xs, y, label=lab, alpha=0.8)
            self._lines.append(ln)
            (mk,) = self.ax.plot([], [], marker="v", linestyle="None")
            self._markers.append(mk)
        self.ax.set_title(self.title)
        self.ax.set_xlabel("Pixels")
        self.ax.set_ylabel("Normalized Intensity")

        self.ax.legend()
        self._vline = self.ax.axvline(self._xs[0] if len(self._xs) else 0, linestyle="--", linewidth=1)
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)

    def _on_motion(self, event):
        if event.inaxes != self.ax or self._xs is None or self._ys_list is None:
            return
        x = event.xdata
        if x is None:
            return
        idx = int(np.clip(np.searchsorted(self._xs, x), 0, len(self._xs) - 1))
        xv = self._xs[idx]
        if self._vline:
            self._vline.set_xdata([xv, xv])
        for y, mk in zip(self._ys_list, self._markers, strict=False):
            if 0 <= idx < len(y):
                mk.set_data([xv], [y[idx]])
        _debounce(self, "_draw_timer", self.draw_idle, ms=40)
        self.hover_index_changed.emit(idx)


# =========================================================
#                   MAIN ANALYSIS WIDGET
# =========================================================
class AnalysisStep(QWidget):
    def __init__(self, data_path=None):
        super().__init__()
        self.working_dir = os.environ.get("MAPEX_WORKDIR", "")
        self.data_list = []
        self.image_list = []
        self.selected_data = None
        self.selected_names = None
        self.latest_roi = None
        self.latest_line = None
        self.data_path = data_path or "data.h5"
        if not os.path.isabs(self.data_path):
            # resolve relative paths into the working folder chosen in main.py
            self.data_path = os.path.abspath(
                os.path.join(
                    self.working_dir or os.getcwd(),
                    os.path.basename(self.data_path),
                )
            )
        self._ternary_active = False
        self._ternary_dlg = None
        self._line_coords = None  # list of (x,y) along current line
        # top-left of current ROI for scatter -> full coords
        self._roi_origin = (0, 0)
        # global toggle for maps/images (not scatter)
        self._highlights_enabled = True
        self._scatter_lasso_enabled = False
        self._ternary_active = False
        self._ternary_lasso_enabled = True
        self._build_ui()

        self.loader = DataLoader(data_path=self.data_path)
        self.loader.data_loaded.connect(self._on_data_loaded)
        self.loader.start()

        self.selected_pixels = []
        self.scatter.selection_made.connect(self._on_scatter_selection)
        self.linescan.hover_index_changed.connect(self._on_linescan_hover)
        self.table_widget = None
        self._point_click_cids = []
        self.point_dialog = None
        self.point_table_dialog = None
        self.selection = SelectionManager(self)
        self.selection.selection_changed.connect(self._on_selection_changed)

        # cache: selected points’ table data
        self._point_rows = []  # list of dicts per point
        self._point_table = None  # dialog (created on demand)

    # ---------- UI ----------
    def _build_ui(self):
        root = QHBoxLayout(self)

        # Left controls
        left = QVBoxLayout()
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        self.log.setFixedWidth(220)
        left.addWidget(self.log)

        self.radio_group = QButtonGroup(self)
        self.rb_corr = QRadioButton("Correlation plot")
        self.rb_line = QRadioButton("Line scan")
        self.rb_point = QRadioButton("Point Data")
        self.rb_corr.setChecked(True)
        for rb in (self.rb_corr, self.rb_line, self.rb_point):
            self.radio_group.addButton(rb)
            left.addWidget(rb)

        # Dropdowns
        self.dropdown_menus = []
        for i in range(3):
            left.addWidget(QLabel(f"Elements {i + 1}"))
            cb = QComboBox()
            cb.currentIndexChanged.connect(self._on_dropdown_changed)
            self.dropdown_menus.append(cb)
            left.addWidget(cb)

        # Brightness sliders
        self.sr = QSlider(Qt.Horizontal)
        self._prep_slider(self.sr)
        self.sg = QSlider(Qt.Horizontal)
        self._prep_slider(self.sg)
        self.sb = QSlider(Qt.Horizontal)
        self._prep_slider(self.sb)
        left.addWidget(QLabel("Element 1 Brightness"))
        left.addWidget(self.sr)
        left.addWidget(QLabel("Element 2 Brightness"))
        left.addWidget(self.sg)
        left.addWidget(QLabel("Element 3 Brightness"))
        left.addWidget(self.sb)
        self.sr.valueChanged.connect(self._refresh_rgb)
        self.sg.valueChanged.connect(self._refresh_rgb)
        self.sb.valueChanged.connect(self._refresh_rgb)

        # Right grid
        self.grid = QGridLayout()
        self.grid.setRowStretch(0, 1)  # top row (maps)
        self.grid.setRowStretch(1, 1)
        for c in range(3):
            self.grid.setColumnStretch(c, 1)
        self.grid.setContentsMargins(0, 0, 0, 0)
        self.grid.setHorizontalSpacing(6)
        self.grid.setVerticalSpacing(6)
        self.map1 = ElementMapView()
        self.map2 = ElementMapView()
        self.rgb = RGBMapView()
        self.grid.addWidget(self.map1, 0, 0)
        self.grid.addWidget(self.map2, 0, 1)
        self.grid.addWidget(self.rgb, 0, 2)

        self.img1 = GrayImageView(title="Image 1")
        self.img2 = GrayImageView(title="Image 2")
        for src in (self.map1, self.map2, self.rgb, self.img1, self.img2):
            src.roi_selected.connect(self._on_roi_selected)
            src.line_selected.connect(self._on_line_selected)

        self.scatter = ScatterView()
        self.linescan = LineScanView()
        self.scatter.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.linescan.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        self.grid.addWidget(self.img1, 1, 0)
        self.grid.addWidget(self.img2, 1, 1)
        # self.grid.addWidget(self.scatter, 1, 2)
        self.cell6 = QWidget()
        self.cell6_layout = QVBoxLayout(self.cell6)
        self.cell6_layout.setContentsMargins(0, 0, 0, 0)
        self.cell6_layout.addWidget(self.scatter)
        self.cell6.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.grid.addWidget(self.cell6, 1, 2)

        root.addLayout(left)
        root.addLayout(self.grid)

        # mode switching
        self.rb_corr.toggled.connect(self._enter_correlation_mode)
        self.rb_line.toggled.connect(self._enter_linescan_mode)
        self.rb_point.toggled.connect(self._enter_pointdata_mode)

        # context menus
        for w in (
            self.map1,
            self.map2,
            self.rgb,
            self.img1,
            self.img2,
            self.scatter,
            self.linescan,
        ):
            w.setContextMenuPolicy(Qt.CustomContextMenu)
            w.customContextMenuRequested.connect(lambda pos, ww=w: self._show_panel_menu(ww, pos))

        self.setContextMenuPolicy(Qt.CustomContextMenu)
        self.customContextMenuRequested.connect(self._show_global_menu)

        for src in (self.map1, self.map2, self.rgb):
            src.roi_selected.connect(self._on_roi_selected)
            src.line_selected.connect(self._on_line_selected)

        self.setLayout(root)

    def _on_ternary_lasso_pixels(self, full_coords):
        """Convert ternary lasso pixels to indices and broadcast
        selection."""
        if not full_coords:
            if hasattr(self, "selection"):
                self.selection.set(set())
            self.selected_pixels = []
            self._apply_highlights_everywhere()
            return

        self.selected_pixels = [(int(x), int(y)) for (x, y) in full_coords]

        idxs = set()
        if hasattr(self, "maps_shape"):
            h, w = self.maps_shape
            for x, y in self.selected_pixels:
                if 0 <= x < w and 0 <= y < h:
                    idxs.add(y * w + x)

        if hasattr(self, "selection"):
            self.selection.set(idxs)  # triggers _on_selection_changed → highlights maps + correlation
        else:
            self._apply_highlights_everywhere()

        # (optional) if Point Data mode is on, refresh the single-pixel table with the first point:
        if getattr(self, "rb_point", None) and self.rb_point.isChecked():
            try:
                x0, y0 = self.selected_pixels[0]
                self._update_point_table_for(x0, y0)
            except Exception:
                pass

    def _on_selection_changed(self, indices: set):
        """Called when SelectionManager updates the selected flat
        indices.

        Keeps highlights, tables, and menus in sync.
        """
        try:
            self._highlight_indices(indices)
            self._selection_has_points = bool(indices)
            # highlight every view that supports it
            self._highlight_indices(indices)

            # If you implemented the per-index metrics table builder, call it
            if hasattr(self, "_compute_point_metrics"):
                try:
                    self._compute_point_metrics(sorted(indices))
                except Exception:
                    pass

            # Update correlation plot highlighting if available
            self._update_correlation_for_selection(indices)

        except Exception as e:
            if hasattr(self, "log"):
                self.log.append(f"[selection] error: {e}")

        tern = getattr(self, "_ternary_dlg", None)
        if tern is not None:
            tern.highlight_from_fullimg_indices(indices, parent=self)

    # def _apply_highlights_everywhere(self):
    #     """Convenience when you have pixel coords (x,y) in
    #     self.selected_pixels.

    #     Uses your existing _highlight_pixels_xy(...) if present.
    #     """
    #     pts = getattr(self, "selected_pixels", []) or []
    #     if hasattr(self, "_highlight_pixels_xy"):
    #         try:
    #             # your code accepts list of (x,y)
    #             self._highlight_pixels_xy(pts)
    #         except Exception:
    #             pass
    #     else:
    #         # Fallback: derive flat indices if you have a shape mapper
    #         if hasattr(self, "maps_shape"):
    #             h, w = self.maps_shape
    #             idxs = set()
    #             for x, y in pts:
    #                 if 0 <= x < w and 0 <= y < h:
    #                     idxs.add(y * w + x)
    #             self._highlight_indices(idxs)

    def _highlight_indices(self, indices: set):
        """Broadcast selected indices to maps/images and the correlation
        plot."""
        # maps/images
        full_shape = getattr(self, "maps_shape", (None, None))
        for name in ("img1", "img2", "map1", "map2", "map3"):
            v = getattr(self, name, None)
            if v is None:
                continue
            if hasattr(v, "set_highlight_indices"):
                try:
                    v.set_highlight_indices(indices, full_shape=full_shape)
                except Exception:
                    pass
            elif hasattr(v, "set_highlight_points"):
                # fallback: convert here
                h, w = full_shape if full_shape else (None, None)
                if h and w:
                    pts = [(int(i) % w, int(i) // w) for i in indices if 0 <= int(i) < h * w]
                    try:
                        v.set_highlight_points(pts)
                    except Exception:
                        pass

        # correlation scatter (index space)
        sc = getattr(self, "scatter", None)
        if sc is not None:
            if hasattr(sc, "set_highlight_indices"):
                try:
                    sc.set_highlight_indices(indices)
                except Exception:
                    pass
            elif hasattr(sc, "update_highlight"):
                try:
                    sc.update_highlight(indices)
                except Exception:
                    pass

    def _update_correlation_for_selection(self, indices: set):
        """Notify the correlation scatter about which points are
        selected.

        Your ScatterView can implement
        update_highlight()/set_highlight_indices().
        """
        sc = getattr(self, "scatter", None)
        if sc is None:
            return

        # prefer an explicit API if present
        if hasattr(sc, "update_highlight") and callable(sc.update_highlight):
            try:
                sc.update_highlight(indices)
                return
            except Exception:
                pass

        if hasattr(sc, "set_highlight_indices") and callable(sc.set_highlight_indices):
            try:
                sc.set_highlight_indices(indices)
            except Exception:
                pass

        if hasattr(sc, "update"):
            try:
                sc.update()
            except Exception:
                pass

    def _open_pointdata_dialog(self):
        # create dialog if needed
        if self.point_dialog is None:
            self.point_dialog = QDialog(self)
            self.point_dialog.setWindowTitle("Point Data")
            self.point_dialog.setMinimumSize(520, 360)

            lay = QVBoxLayout(self.point_dialog)

            # build a dedicated table for the dialog (independent of cell6)
            self.point_table_dialog = QTableWidget()
            self.point_table_dialog.setColumnCount(4)
            self.point_table_dialog.setHorizontalHeaderLabels(
                [
                    "Element",
                    "Intensity",
                    "Quantification",
                    "Quantification (Normalized)",
                ]
            )
            hdr = self.point_table_dialog.horizontalHeader()
            hdr.setStretchLastSection(True)
            self.point_table_dialog.verticalHeader().setVisible(False)
            self.point_table_dialog.setShowGrid(True)

            # fill all channel names once
            if getattr(self, "data_list", None):
                nrows = len(self.data_list)
                self.point_table_dialog.setRowCount(nrows)
                for i, (nm, _) in enumerate(self.data_list):
                    self.point_table_dialog.setItem(i, 0, QTableWidgetItem(self._display_name(nm)))
                    self.point_table_dialog.setItem(i, 1, QTableWidgetItem("--"))
                    self.point_table_dialog.setItem(i, 2, QTableWidgetItem("--"))
                    self.point_table_dialog.setItem(i, 3, QTableWidgetItem("--"))

            # close button
            btn = QPushButton("Close")
            btn.clicked.connect(self.point_dialog.close)

            lay.addWidget(self.point_table_dialog)
            lay.addWidget(btn)

        self.point_dialog.show()
        self.point_dialog.raise_()
        self.point_dialog.activateWindow()

    def _show_in_cell6(self, widget):
        # remove everything currently in the 6th slot and show `widget`
        while self.cell6_layout.count():
            itm = self.cell6_layout.takeAt(0)
            w = itm.widget()
            if w:
                w.setParent(None)
        self.cell6_layout.addWidget(widget)

    def _open_ternary_dialog(self):
        """Open (or re-open) the triangular/ternary dialog and populate
        it from the CURRENT VIEW using
        _update_ternary_with_current_view().

        No calls to _get_current_channels().
        """
        # ensure flags exist
        if not hasattr(self, "_ternary_lasso_enabled"):
            self._ternary_lasso_enabled = True

        # create dialog if needed
        if getattr(self, "_ternary_dlg", None) is None:
            try:
                self._ternary_dlg = TernaryDialog(self)
                try:
                    self._ternary_dlg.lasso_pixels_xy.connect(self._on_ternary_lasso_pixels)
                except Exception:
                    pass
            except Exception as e:
                print("Failed to create TernaryDialog:", e)
                self._ternary_dlg = None
                return

        # set labels if available (first three symbols)
        if getattr(self, "selected_symbols", None):
            try:
                self._ternary_dlg.set_labels(self.selected_symbols[:3])
            except Exception:
                pass

        # populate from current visible data/ROI
        if hasattr(self, "_update_ternary_with_current_view"):
            self._update_ternary_with_current_view()

        # match lasso toggle state
        if hasattr(self._ternary_dlg, "enable_lasso"):
            self._ternary_dlg.enable_lasso(self._ternary_lasso_enabled)

        # show/raise dialog
        try:
            self._ternary_dlg.show()
            self._ternary_dlg.raise_()
        except Exception:
            pass

        # keep ternary synced with pan/zoom of the scatter (connect only once)
        if not hasattr(self, "_ternary_callbacks_connected") or not self._ternary_callbacks_connected:
            if hasattr(self, "scatter") and hasattr(self.scatter, "ax") and self.scatter.ax is not None:
                try:
                    self.scatter.ax.callbacks.connect(
                        "xlim_changed",
                        lambda ax: self._update_ternary_with_current_view(),
                    )
                    self.scatter.ax.callbacks.connect(
                        "ylim_changed",
                        lambda ax: self._update_ternary_with_current_view(),
                    )
                    self._ternary_callbacks_connected = True
                except Exception:
                    self._ternary_callbacks_connected = True  # avoid retry loops

    # ---------- Highlight helpers ----------
    def _apply_highlights_everywhere(self):
        targets = (self.map1, self.map2, self.rgb, self.img1, self.img2)
        for w in targets:
            if hasattr(w, "set_highlight_points"):
                w.set_highlight_points(self.selected_pixels)
            if hasattr(w, "set_highlight_visible"):
                w.set_highlight_visible(self._highlights_enabled)

    def _clear_highlights(self):
        self.selected_pixels = []
        """Set data path."""
        targets = (self.map1, self.map2, self.rgb, self.img1, self.img2)
        for w in targets:
            if hasattr(w, "set_highlight_points"):
                w.set_highlight_points([])

    def _update_highlight_visibility_only(self):
        targets = (self.map1, self.map2, self.rgb, self.img1, self.img2)
        for w in targets:
            if hasattr(w, "set_highlight_visible"):
                w.set_highlight_visible(self._highlights_enabled)

    def _clear_all_cursors(self):
        for w in (self.map1, self.map2, self.rgb, self.img1, self.img2):
            if hasattr(w, "clear_cursor"):
                w.clear_cursor()

    def _highlight_pixels_xy(self, coords):
        """coords: iterable of (x, y) in FULL-IMAGE pixel coordinates (int).
        Sets the highlight mask and refreshes overlays.
        """
        # infer image size from your first selected channel
        if not getattr(self, "selected_data", None):
            return
        h, w = self.selected_data[0].shape[:2]

        # ensure mask exists and correct shape
        if not hasattr(self, "_highlight_mask") or self._highlight_mask.shape != (h, w):
            self._highlight_mask = np.zeros((h, w), dtype=bool)

        # set pixels
        for x, y in coords:
            if 0 <= x < w and 0 <= y < h:
                self._highlight_mask[y, x] = True

        # show them (guard if method missing)
        if hasattr(self, "_refresh_highlights"):
            self._refresh_highlights()

    # ---------- Events from children ----------
    def _on_scatter_selection(self, coords_xy_list):

        ox, oy = self._roi_origin
        self.selected_pixels = [(x + ox, y + oy) for (x, y) in (coords_xy_list or [])]
        self._apply_highlights_everywhere()

    def _fix_roi_bounds(self, roi, W, H, min_size=2):
        x1, y1, x2, y2 = roi
        x1, x2 = sorted((int(x1), int(x2)))
        y1, y2 = sorted((int(y1), int(y2)))
        x1 = max(0, min(x1, W - 1))
        x2 = max(0, min(x2, W))
        y1 = max(0, min(y1, H - 1))
        y2 = max(0, min(y2, H))
        if x2 - x1 < min_size:
            cx = (x1 + x2) // 2
            x1 = max(0, cx - min_size // 2)
            x2 = min(W, x1 + min_size)
        if y2 - y1 < min_size:
            cy = (y1 + y2) // 2
            y1 = max(0, cy - min_size // 2)
            y2 = min(H, y1 + min_size)
        return x1, y1, x2, y2

    def _prep_slider(self, sld):
        sld.setRange(0, 200)
        sld.setValue(100)
        sld.setFixedWidth(200)

    # ---------- Data ----------
    def _on_data_loaded(self, data_list, image_list):
        self.data_list = data_list or []
        self.image_list = image_list or []
        if self.data_list:
            self.log.append(f"Loaded {len(self.data_list)} map(s); " f"{len(self.image_list)} image(s).")
            for cb in self.dropdown_menus:
                cb.clear()
                cb.addItems([name for name, _ in self.data_list])
            # ensure a valid selection exists
            for cb in self.dropdown_menus:
                if cb.count() > 0 and cb.currentIndex() < 0:
                    cb.setCurrentIndex(0)
            self._refresh_views_from_dropdowns()
        else:
            self.log.append(f"Failed to load from {self.data_path}. No maps found.")

    def _wd(self, *parts):
        base = self.working_dir or os.getcwd()
        return os.path.abspath(os.path.join(base, *parts))


    def _on_dropdown_changed(self, _):
        self._refresh_views_from_dropdowns()

    def _refresh_views_from_dropdowns(self):
        """Refresh maps, images, RGB composite, and correlation/line
        views based on dropdown selection.

        Also pushes per-pixel σ to the scatter for uncertainty-aware
        stats (MC/weighted).
        """
        if not getattr(self, "dropdown_menus", None) or len(self.dropdown_menus) < 3:
            return
        if not getattr(self, "data_list", None):
            return

        def base_name(raw: str) -> str:
            return os.path.splitext(raw)[0].strip()

        def full_name(raw: str) -> str:
            return f"{base_name(raw).capitalize()} Map"

        total = len(self.data_list)
        idxs = []
        for cb in self.dropdown_menus[:3]:
            i = cb.currentIndex()
            idxs.append(0 if i is None or i < 0 or i >= total else i)

        maps = [self.data_list[i][1] for i in idxs]
        raw_names = [self.data_list[i][0] for i in idxs]
        names_full = [full_name(n) for n in raw_names]
        names_short = [base_name(n).capitalize() for n in raw_names]

        self.selected_data = maps
        if self.selected_data and self.selected_data[0] is not None:
            self.maps_shape = self.selected_data[0].shape[:2]
        self.selected_names = names_full
        self.selected_symbols = names_short

        # FULL VIEW -> origin reset
        self._roi_origin = (0, 0)

        br = self.sr.value() / 100.0 if hasattr(self, "sr") else 1.0
        bg = self.sg.value() / 100.0 if hasattr(self, "sg") else 1.0
        bb = self.sb.value() / 100.0 if hasattr(self, "sb") else 1.0

        # ---- Update individual map views ------------------------------------
        try:
            self.map1.set_data(
                maps[0],
                cmap="Reds",
                title=self.selected_names[0],
                offset=(0, 0),
            )
        except TypeError:
            self.map1.set_data(maps[0], cmap="Reds", title=self.selected_names[0])

        try:
            self.map2.set_data(
                maps[1],
                cmap="Greens",
                title=self.selected_names[1],
                offset=(0, 0),
            )
        except TypeError:
            self.map2.set_data(maps[1], cmap="Greens", title=self.selected_names[1])

        # ---- Update RGB composite -------------------------------------------
        try:
            self.rgb.set_data(
                maps[0],
                maps[1],
                maps[2],
                names=self.selected_symbols,
                br=br,
                bg=bg,
                bb=bb,
                offset=(0, 0),
            )
        except TypeError:
            self.rgb.set_data(
                maps[0],
                maps[1],
                maps[2],
                names=self.selected_symbols,
                br=br,
                bg=bg,
                bb=bb,
            )

        # ---- Update images if any -------------------------------------------
        if getattr(self, "image_list", None):
            try:
                self.img1.set_data(self.image_list[0][1], title="Image 1", offset=(0, 0))
            except TypeError:
                self.img1.set_data(self.image_list[0][1], title="Image 1")
            if len(self.image_list) > 1:
                try:
                    self.img2.set_data(self.image_list[1][1], title="Image 2", offset=(0, 0))
                except TypeError:
                    self.img2.set_data(self.image_list[1][1], title="Image 2")

        # ---- Correlation or Line views --------------------------------------
        do_corr = getattr(self, "rb_corr", None) and self.rb_corr.isChecked()
        do_line = getattr(self, "rb_line", None) and self.rb_line.isChecked()

        if do_corr and hasattr(self, "scatter"):
            # Set scatter data
            self.scatter.set_data(
                maps[0],
                maps[1],
                f"{self.selected_names[0]} Intensity",
                f"{self.selected_names[1]} Intensity",
                remember_limits=False,
            )

            # Push per-pixel σ for uncertainty-aware stats
            # Prefer project helper if available; else basic Poisson
            # sqrt(counts)
            try:
                sx = _estimate_sigma_from_counts(maps[0], scale=1.0)  # provided in helpers
                sy = _estimate_sigma_from_counts(maps[1], scale=1.0)
            except Exception:
                a0 = np.asarray(maps[0], dtype=float)
                a0[a0 < 0] = 0.0
                a1 = np.asarray(maps[1], dtype=float)
                a1[a1 < 0] = 0.0
                sx = np.sqrt(a0)
                sy = np.sqrt(a1)

            if hasattr(self.scatter, "set_uncertainty"):
                self.scatter.set_uncertainty(sx, sy)

            # Respect last user choice, else defaults
            mc_enabled = getattr(
                self.scatter,
                "_mc_enabled",
                getattr(self, "_mc_enabled_default", False),
            )
            mc_iters = getattr(
                self.scatter,
                "_mc_iters",
                getattr(self, "_mc_iters_default", 1000),
            )
            if hasattr(self.scatter, "enable_mc"):
                self.scatter.enable_mc(mc_enabled, iters=mc_iters)

        elif do_line and hasattr(self, "_plot_center_line"):
            self._plot_center_line()

        # ---- ROI enablement -------------------------------------------------
        enable_roi = bool(do_corr)
        for w in (self.map1, self.map2, self.rgb, self.img1, self.img2):
            if hasattr(w, "set_roi_enabled"):
                w.set_roi_enabled(enable_roi)

        # ---- apply highlights and clean cursors -----------------------------
        if hasattr(self, "_apply_highlights_everywhere"):
            self._apply_highlights_everywhere()
        if hasattr(self, "_clear_all_cursors"):
            self._clear_all_cursors()

        # ---- update ternary if present --------------------------------------
        if hasattr(self, "_update_ternary_with_current_view"):
            self._update_ternary_with_current_view()

    # ---------- Modes ----------
    def _enter_correlation_mode(self, checked):
        if not checked:
            return
        self.log.append("Switching to Correlation plot mode...")
        self._refresh_views_from_dropdowns()
        self._swap_bottom_right(to="scatter")
        for w in (self.map1, self.map2, self.rgb, self.img1, self.img2):
            w.set_roi_enabled(True)

    def _enter_linescan_mode(self, checked):
        if not checked:
            return
        self.log.append("Switching to Line Scan mode...")
        self._swap_bottom_right(to="linescan")
        self.map1.set_roi_enabled(False)
        self.map2.set_roi_enabled(False)
        self.rgb.set_roi_enabled(False)

    def _enter_pointdata_mode(self, checked):
        if not checked:
            # leaving Point Data -> stop per-pixel picking
            self._disconnect_point_clicks()
            # restore bottom-right to scatter (if you changed it anywhere)
            try:
                self._show_in_cell6(self.scatter)
            except Exception:
                pass
            return

        # entering Point Data: open popup table (no splitter in the grid)
        self.log.append("Switching to Point Data mode...")
        self._open_pointdata_dialog()  # creates dialog + table if needed
        self._connect_point_clicks()  # enable per-pixel picking across maps/images

    def _ensure_point_table(self):
        # We want ALL channels from the dataset, not just the 3 selected
        # Use self.data_list: [(name, array), ...]
        if not getattr(self, "data_list", None):
            return

        names_all = [self._display_name(nm) for (nm, _) in self.data_list]
        nrows = len(names_all)

        if self.table_widget is None:
            self.table_widget = QTableWidget()
            hdr = self.table_widget.horizontalHeader()
            hdr.setStretchLastSection(True)
            self.table_widget.verticalHeader().setVisible(False)
            self.table_widget.setShowGrid(True)
            # compact table height (works with the splitter)
            row_h = 22
            self.table_widget.verticalHeader().setDefaultSectionSize(row_h)
            max_visible = min(nrows, 8)
            cap_h = (max_visible * row_h) + 40
            self.table_widget.setMinimumHeight(min(cap_h, 260))
            self.table_widget.setMaximumHeight(260)

            from PyQt5.QtWidgets import QSizePolicy

            self.table_widget.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.MinimumExpanding)

            self.table_widget.setColumnCount(4)
            self.table_widget.setHorizontalHeaderLabels(
                [
                    "Element",
                    "Intensity",
                    "Quantification",
                    "Quantification (Normalized)",
                ]
            )

        # ensure rows match channel count
        if self.table_widget.rowCount() != nrows:
            self.table_widget.setRowCount(nrows)

        # fill the first column with all element/map names
        for i, nm in enumerate(names_all):
            itm = self.table_widget.item(i, 0)
            if itm is None or itm.text() != nm:
                self.table_widget.setItem(i, 0, QTableWidgetItem(nm))
        # clear the numeric columns (fresh fill on click)
        for i in range(nrows):
            self.table_widget.setItem(i, 1, QTableWidgetItem("--"))
            self.table_widget.setItem(i, 2, QTableWidgetItem("--"))
            self.table_widget.setItem(i, 3, QTableWidgetItem("--"))

    def _load_calibration_data(self):
        """Returns dict like {'FE': (m, c), 'SI': (m, c), ...} from calibration_data.xlsx
        Column order expected: Element | m | c
        """
        path = "calibration_data.xlsx"  # ← per your note
        calib = {}
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            sh = wb.active
            # try to detect header
            start_row = 2
            hdr = [v if v is not None else "" for v in next(sh.iter_rows(min_row=1, max_row=1, values_only=True))]
            if any(s for s in hdr):
                start_row = 2
            else:
                start_row = 1

            for row in sh.iter_rows(min_row=start_row, values_only=True):
                if not row:
                    continue
                element = row[0]
                if element is None:
                    continue
                m = row[1] if len(row) > 1 else None
                c = row[2] if len(row) > 2 else None
                key = self._norm_elem_key(str(element))
                if key:
                    try:
                        m = float(m)
                    except Exception:
                        m = None
                    try:
                        c = float(c)
                    except Exception:
                        c = 0.0
                    calib[key] = (m, c)
            if calib:
                self.log.append(f"Loaded calibration data from {path} ({len(calib)} entries).")
            else:
                self.log.append(f"No usable calibration rows found in {path}.")
        except Exception as e:
            self.log.append(f"Failed to load {path}: {e}")
        return calib

    def _norm_elem_key(self, name: str):
        # strip extension, content in parentheses/brackets, "intensity" suffix
        # etc.
        s = name.strip()
        s = re.sub(r"\.[A-Za-z0-9]+$", "", s)  # remove .txt/.csv/etc.
        s = re.sub(r"\s*\(.*?\)\s*", "", s)  # remove (...) parts
        s = re.sub(r"\s*\[.*?\]\s*", "", s)  # remove [...] parts
        s = s.replace("intensity", "").replace("Intensity", "")
        s = re.sub(r"[^A-Za-z]+", "", s)  # keep letters only
        return s.upper() if s else None

    def _set_point_row_values(self, table, row, intensity, qv, qnorm):
        if table is None:
            return
        if row >= table.rowCount():
            table.setRowCount(row + 1)
        if intensity is not None:
            table.setItem(row, 1, QTableWidgetItem(f"{intensity:.2f}"))
        if qv is None:
            table.setItem(row, 2, QTableWidgetItem(""))
        elif qv == "Undefined":
            table.setItem(row, 2, QTableWidgetItem("Undefined"))
        else:
            table.setItem(row, 2, QTableWidgetItem(f"{qv:.2f}"))
        if qnorm is None:
            table.setItem(row, 3, QTableWidgetItem(""))
        else:
            table.setItem(row, 3, QTableWidgetItem(f"{qnorm:.2f}%"))

    def _update_point_table_for(self, x_full, y_full):
        if not getattr(self, "data_list", None):
            return

        calib = self._load_calibration_data()
        quant_vals = []
        intensities = []

        # compute per-channel
        for i, (raw_name, arr) in enumerate(self.data_list):
            if arr is None or arr.size == 0:
                intensities.append(None)
                quant_vals.append(0.0)
                continue
            h, w = arr.shape[:2]
            if not (0 <= x_full < w and 0 <= y_full < h):
                intensities.append(None)
                quant_vals.append(0.0)
                continue
            intensity = float(arr[y_full, x_full])
            intensities.append(intensity)

            key = self._norm_elem_key(str(raw_name))
            if key in calib:
                m, c = calib[key]
                if m is not None and m != 0:
                    qv = (intensity - (c or 0.0)) / m
                    if qv < 0:
                        qv = 0.0
                    quant_vals.append(qv)
                else:
                    quant_vals.append(0.0)
            else:
                quant_vals.append(0.0)

        total = sum(q for q in quant_vals if q > 0) or 1.0

        # write to both tables (if present)
        for i, qv in enumerate(quant_vals):
            intensity = intensities[i]
            qnorm = (qv / total) * 100.0 if qv > 0 else None

            # main (old cell6) table if you kept it
            self._set_point_row_values(
                getattr(self, "table_widget", None),
                i,
                intensity,
                (qv if qv > 0 else None),
                qnorm,
            )
            # popup dialog table
            self._set_point_row_values(
                self.point_table_dialog,
                i,
                intensity,
                (qv if qv > 0 else None),
                qnorm,
            )

        # ensure repaint
        if self.point_table_dialog is not None:
            self.point_table_dialog.viewport().update()

    def _connect_point_clicks(self):
        self._disconnect_point_clicks()

        def _mk(handler_widget):
            def _on_click(ev):
                if not self.rb_point.isChecked():
                    return
                if ev.inaxes is None or ev.xdata is None or ev.ydata is None:
                    return
                # full-image coordinates: account for panel’s offset
                ox, oy = getattr(handler_widget, "offset", (0, 0))
                x_full = int(ev.xdata) + int(ox)
                y_full = int(ev.ydata) + int(oy)

                # 1) Update popup table (per-channel) for this pixel
                self._update_point_table_for(x_full, y_full)

                # 2) Make this a live selection so highlights & menus respond
                self.selected_pixels = [(x_full, y_full)]
                self._apply_highlights_everywhere()

                self.log.append(f"Selected pixel: ({x_full}, {y_full})")

            return _on_click

        for w in (self.map1, self.map2, self.rgb, self.img1, self.img2):
            cid = w.figure.canvas.mpl_connect("button_press_event", _mk(w))
            self._point_click_cids.append((w, cid))

    def _disconnect_point_clicks(self):
        for w, cid in self._point_click_cids:
            try:
                w.figure.canvas.mpl_disconnect(cid)
            except Exception:
                pass
        self._point_click_cids = []

    def _swap_bottom_right(self, to="scatter"):
        # Clear current widget in cell6
        while self.cell6_layout.count():
            itm = self.cell6_layout.takeAt(0)
            w = itm.widget()
            if w:
                w.setParent(None)

        if to == "scatter":
            self.scatter.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self.cell6_layout.addWidget(self.scatter)
            if self.selected_data and self.selected_names:
                self.scatter.set_data(
                    self.selected_data[0],
                    self.selected_data[1],
                    f"{self.selected_names[0]} Intensity",
                    f"{self.selected_names[1]} Intensity",
                    remember_limits=False,
                )
        else:
            self.linescan.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            self.cell6_layout.addWidget(self.linescan)
            if self.selected_data and self.selected_names:
                self._plot_center_line()

    # ---------- Context menus ----------
    def _show_global_menu(self, pos):
        menu = QMenu(self)
        act_quick = menu.addAction("Quick Reset")
        act_refresh = menu.addAction("Refresh")
        act_toggle = menu.addAction("Highlight Overlay")
        act_clear = menu.addAction("Clear Highlights")

        act_point_table = None
        if getattr(self, "selected_pixels", None):
            if len(self.selected_pixels) > 0:
                act_point_table = menu.addAction("Make Point Data Table")

        chosen = menu.exec_(self.mapToGlobal(pos))
        if chosen == act_quick:
            self._quick_reset()
        elif chosen == act_refresh:
            self._full_reset()
        elif chosen == act_toggle:
            self._highlights_enabled = not self._highlights_enabled
            self._update_highlight_visibility_only()
        elif chosen == act_clear:
            self._clear_highlights()
        elif (act_point_table is not None) and (chosen is act_point_table):
            self._open_selected_points_table_dialog()

    def _compute_quant_for_xy(self, x_full, y_full):
        """Compute per-channel Intensity, Quant, NormQuant for a single
        pixel (x_full,y_full).

        Returns: (display_names, intensities, quants, norm_quants)
        """
        if not getattr(self, "data_list", None):
            return [], [], [], []

        calib = self._load_calibration_data()
        intensities, quants = [], []
        names = [self._display_name(nm) for (nm, _) in self.data_list]

        # gather intensities + raw quants
        for raw_name, arr in self.data_list:
            if arr is None or arr.size == 0:
                intensities.append(None)
                quants.append(0.0)
                continue
            h, w = arr.shape[:2]
            if not (0 <= x_full < w and 0 <= y_full < h):
                intensities.append(None)
                quants.append(0.0)
                continue
            intensity = float(arr[y_full, x_full])
            intensities.append(intensity)

            key = self._norm_elem_key(str(raw_name))
            if key in calib:
                m, c = calib[key]
                if m is not None and m != 0:
                    qv = (intensity - (c or 0.0)) / m
                    if qv < 0:
                        qv = 0.0
                    quants.append(qv)
                else:
                    quants.append(0.0)
            else:
                quants.append(0.0)

        total = sum(q for q in quants if q > 0) or 1.0
        nquants = [(q / total) * 100.0 if q > 0 else None for q in quants]
        return names, intensities, quants, nquants

    def _open_selected_points_table_dialog(self):
        """Popup table with one row per selected pixel.

        Columns: x, y, then Q:<elem> and NQ%:<elem>.
        Clicking a row re-selects/focuses that pixel everywhere.
        """
        if not getattr(self, "selected_pixels", None) or len(self.selected_pixels) == 0:
            QMessageBox.information(
                self,
                "No selection",
                "Select pixels first (click a map or lasso).",
            )
            return

        # Build column headers (from dataset channels)
        if not getattr(self, "data_list", None):
            QMessageBox.warning(self, "No data", "No maps loaded.")
            return
        elem_names = [self._display_name(nm) for (nm, _) in self.data_list]
        q_headers = [f"Q:{e}" for e in elem_names]
        nq_headers = [f"NQ%:{e}" for e in elem_names]

        from PyQt5.QtWidgets import (
            QAbstractItemView,
            QDialog,
            QDialogButtonBox,
            QTableWidget,
            QTableWidgetItem,
            QVBoxLayout,
        )

        dlg = QDialog(self)
        dlg.setWindowTitle("Selected Points – Quantification")
        lay = QVBoxLayout(dlg)

        tbl = QTableWidget()
        tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
        tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)

        headers = ["x", "y"] + q_headers + nq_headers
        tbl.setColumnCount(len(headers))
        tbl.setHorizontalHeaderLabels(headers)
        tbl.verticalHeader().setVisible(False)

        # Fill rows
        tbl.setRowCount(len(self.selected_pixels))
        for r, (x, y) in enumerate(self.selected_pixels):
            names, ints_, quants, nquants = self._compute_quant_for_xy(x, y)
            # x, y
            tbl.setItem(r, 0, QTableWidgetItem(str(x)))
            tbl.setItem(r, 1, QTableWidgetItem(str(y)))
            # Q:<elem>
            for i, qv in enumerate(quants):
                it = QTableWidgetItem("" if qv is None else f"{qv:.6g}")
                it.setTextAlignment(Qt.AlignCenter)
                tbl.setItem(r, 2 + i, it)
            # NQ%:<elem>
            base = 2 + len(quants)
            for i, nv in enumerate(nquants):
                it = QTableWidgetItem("" if nv is None else f"{nv:.4g}")
                it.setTextAlignment(Qt.AlignCenter)
                tbl.setItem(r, base + i, it)

        tbl.resizeColumnsToContents()
        tbl.horizontalHeader().setStretchLastSection(True)

        # Selecting a row focuses that pixel everywhere
        def _on_sel_changed():
            items = tbl.selectedItems()
            if not items:
                return
            row = tbl.currentRow()
            try:
                x = int(tbl.item(row, 0).text())
                y = int(tbl.item(row, 1).text())
            except Exception:
                return
            self.selected_pixels = [(x, y)]
            self._apply_highlights_everywhere()

        tbl.itemSelectionChanged.connect(_on_sel_changed)

        lay.addWidget(tbl)
        btns = QDialogButtonBox(QDialogButtonBox.Close, parent=dlg)
        btns.rejected.connect(dlg.close)
        lay.addWidget(btns)

        dlg.setMinimumSize(640, 420)
        dlg.show()
        dlg.raise_()
        dlg.activateWindow()

    def _show_panel_menu(self, widget, pos):
        """Context menu for any panel.

        Adds common actions and panel-specific items. Now includes
        Uncertainty Propagation (MC) controls on the correlation panel.
        """
        # ---- ensure state vars exist (idempotent) ----
        if not hasattr(self, "_highlights_enabled"):
            self._highlights_enabled = True
        if not hasattr(self, "_scatter_lasso_enabled"):
            self._scatter_lasso_enabled = False
        if not hasattr(self, "_ternary_active"):
            self._ternary_active = False
        if not hasattr(self, "_ternary_lasso_enabled"):
            self._ternary_lasso_enabled = True
        if not hasattr(self, "_ternary_dlg"):
            self._ternary_dlg = None
        if not hasattr(self, "_mc_enabled_default"):
            self._mc_enabled_default = False
        if not hasattr(self, "_mc_iters_default"):
            self._mc_iters_default = 1000

        menu = QMenu(widget)

        # ---- Common actions -------------------------------------------------
        act_quick = menu.addAction("Quick Reset")
        act_refresh = menu.addAction("Refresh")

        act_toggle_hl = menu.addAction("Highlight Overlay")
        act_toggle_hl.setCheckable(True)
        act_toggle_hl.setChecked(self._highlights_enabled)

        act_clear_hl = menu.addAction("Clear Highlights")
        act_point_table = menu.addAction("Make Point Data Table")

        # placeholders for conditional actions
        act_select_pixels = None
        act_save_corr = None
        act_save_line = None
        act_save_roi = None
        act_ternary = None
        act_ternary_lasso = None

        # New (scatter) uncertainty actions
        act_mc_toggle = None
        act_mc_iters = None
        act_mc_hist = None

        # ---- Panel-specific actions -----------------------------------------
        is_scatter_ctx = (
            (widget is getattr(self, "scatter", None)) and getattr(self, "rb_corr", None) and self.rb_corr.isChecked()
        )
        is_line_ctx = (
            (widget is getattr(self, "linescan", None)) and getattr(self, "rb_line", None) and self.rb_line.isChecked()
        )

        # Correlation scatter panel
        if is_scatter_ctx:
            # Select Pixels (Lasso) — checkable
            act_select_pixels = menu.addAction("Select Pixels (Lasso)")
            act_select_pixels.setCheckable(True)
            act_select_pixels.setChecked(self._scatter_lasso_enabled)

            # Save correlation
            act_save_corr = menu.addAction("Save Correlation (CSV)")

            # Triangular Plot toggle — checkable
            act_ternary = menu.addAction("Triangular Plot")
            act_ternary.setCheckable(True)
            act_ternary.setChecked(self._ternary_active)

            # Ternary Lasso toggle — only if dialog exists and ternary is
            # active
            if self._ternary_active and (self._ternary_dlg is not None):
                act_ternary_lasso = menu.addAction("Ternary Lasso")
                act_ternary_lasso.setCheckable(True)
                act_ternary_lasso.setChecked(self._ternary_lasso_enabled)

            # ---- Uncertainty (MC) controls ----------------------------------
            # Show these only if the scatter supports uncertainty methods
            scat = getattr(self, "scatter", None)
            has_mc = hasattr(scat, "enable_mc")
            if has_mc:
                # Toggle (checkable)
                act_mc_toggle = menu.addAction("Propagate Uncertainty (Monte Carlo)")
                act_mc_toggle.setCheckable(True)
                # Ask the scatter for current state if available; else use
                # defaults
                curr_mc = getattr(scat, "_mc_enabled", self._mc_enabled_default)
                act_mc_toggle.setChecked(bool(curr_mc))

                # Set iterations
                act_mc_iters = menu.addAction(
                    f"Set MC Iterations (current: {getattr(scat, '_mc_iters', self._mc_iters_default)})"
                )
                # Optional histogram save if implemented
                if hasattr(scat, "save_mc_histogram"):
                    act_mc_hist = menu.addAction("Save r (MC) Histogram...")

        # Line-scan panel
        if is_line_ctx:
            act_save_line = menu.addAction("Save Line Scan (CSV)")

        # Any image/ROI panel that's not scatter or linescan
        if widget not in (
            getattr(self, "scatter", None),
            getattr(self, "linescan", None),
        ):
            act_save_roi = menu.addAction("Save ROI")

        # ---- Execute and capture the chosen action --------------------------
        chosen = menu.exec_(widget.mapToGlobal(pos))
        if chosen is None:
            return

        # ---- Handle actions (common) ----------------------------------------
        if chosen is act_quick and hasattr(self, "_quick_reset"):
            self._quick_reset()
            return

        if chosen is act_refresh and hasattr(self, "_refresh_views_from_dropdowns"):
            self._refresh_views_from_dropdowns()
            return

        if chosen is act_toggle_hl and hasattr(self, "_refresh_highlights"):
            self._highlights_enabled = not self._highlights_enabled
            self._refresh_highlights()
            return

        if chosen is act_clear_hl and hasattr(self, "_clear_highlights"):
            self._clear_highlights()
            return

        if "act_point_table" in locals() and (chosen is act_point_table):
            if hasattr(self, "_open_pointdata_dialog"):
                self._open_pointdata_dialog()
            return

        # ---- Handle actions (correlation scatter) ---------------------------
        if (act_select_pixels is not None) and (chosen is act_select_pixels):
            self._scatter_lasso_enabled = not self._scatter_lasso_enabled
            if hasattr(self, "scatter") and hasattr(self.scatter, "enable_lasso"):
                self.scatter.enable_lasso(self._scatter_lasso_enabled)
            elif hasattr(self, "_enable_scatter_lasso"):
                self._enable_scatter_lasso(self._scatter_lasso_enabled)
            return

        if (act_save_corr is not None) and (chosen is act_save_corr):
            if hasattr(self, "scatter") and hasattr(self.scatter, "save_visible_correlation_csv"):
                self.scatter.save_visible_correlation_csv(self)
            elif hasattr(self, "_save_correlation_csv"):
                self._save_correlation_csv()
            elif hasattr(self, "_save_correlation"):
                self._save_correlation()
            return

        if (act_ternary is not None) and (chosen is act_ternary):
            self._ternary_active = not self._ternary_active
            if self._ternary_active:
                if hasattr(self, "_open_ternary_dialog"):
                    self._open_ternary_dialog()
                elif hasattr(self, "_update_ternary_with_current_view"):
                    if self._ternary_dlg is None and hasattr(self, "TernaryDialog"):
                        try:
                            self._ternary_dlg = self.TernaryDialog(self)  # type: ignore
                        except Exception:
                            pass
                    if self._ternary_dlg is not None:
                        self._update_ternary_with_current_view()
                        try:
                            self._ternary_dlg.show()
                        except Exception:
                            pass
            else:
                if self._ternary_dlg is not None:
                    try:
                        self._ternary_dlg.close()
                    except Exception:
                        pass
                    self._ternary_dlg = None
            return

        if (act_ternary_lasso is not None) and (chosen is act_ternary_lasso):
            self._ternary_lasso_enabled = not self._ternary_lasso_enabled
            if (self._ternary_dlg is not None) and hasattr(self._ternary_dlg, "enable_lasso"):
                self._ternary_dlg.enable_lasso(self._ternary_lasso_enabled)
            return

        # ---- Uncertainty (MC) handlers --------------------------------------
        if (act_mc_toggle is not None) and (chosen is act_mc_toggle):
            if hasattr(self, "scatter") and hasattr(self.scatter, "enable_mc"):
                self.scatter.enable_mc(act_mc_toggle.isChecked(), iters=None)
                # remember preference session-wide
                self._mc_enabled_default = act_mc_toggle.isChecked()
            return

        if (act_mc_iters is not None) and (chosen is act_mc_iters):
            try:
                iters, ok = QInputDialog.getInt(
                    self,
                    "MC Iterations",
                    "Number of draws:",
                    getattr(self.scatter, "_mc_iters", self._mc_iters_default),
                    100,
                    50000,
                    100,
                )
                if ok and hasattr(self, "scatter") and hasattr(self.scatter, "enable_mc"):
                    self.scatter.enable_mc(
                        getattr(
                            self.scatter,
                            "_mc_enabled",
                            self._mc_enabled_default,
                        ),
                        iters=int(iters),
                    )
                    self._mc_iters_default = int(iters)
            except Exception as e:
                print("MC Iterations dialog failed:", e)
            return

        if (act_mc_hist is not None) and (chosen is act_mc_hist):
            if hasattr(self, "scatter") and hasattr(self.scatter, "save_mc_histogram"):
                self.scatter.save_mc_histogram(self)
            else:
                try:
                    QMessageBox.information(
                        self,
                        "Unavailable",
                        "MC histogram export is not implemented on the scatter.",
                    )
                except Exception:
                    pass
            return

        # ---- Handle actions (line-scan) -------------------------------------
        if (act_save_line is not None) and (chosen is act_save_line):
            if hasattr(self, "linescan") and hasattr(self.linescan, "save_csv"):
                self.linescan.save_csv(self)
            elif hasattr(self, "_save_linescan_csv"):
                self._save_linescan_csv()
            return

        # ---- Handle actions (ROI) -------------------------------------------
        if (act_save_roi is not None) and (chosen is act_save_roi):
            if hasattr(self, "_save_roi"):
                self._save_roi()
            elif hasattr(self, "save_roi"):
                self.save_roi()
            return

    def _display_name(self, raw_name: str) -> str:
        base = os.path.splitext(raw_name)[0]
        return f"{base.strip().capitalize()} Map"

    # ---------- Interactions ----------
    def _on_roi_selected(self, roi):
        """Zoom all panels to ROI, update scatter with remembered
        limits, and keep highlights."""
        self.latest_roi = roi
        if not self.selected_data or not self.selected_names or not self.image_list:
            return

        m0, m1, m2 = self.selected_data
        h, w = m0.shape[:2]
        x1, y1, x2, y2 = self._fix_roi_bounds(roi, w, h, min_size=2)

        # store ROI origin so lasso coords map to full image
        self._roi_origin = (x1, y1)

        br = self.sr.value() / 100.0
        bg = self.sg.value() / 100.0
        bb = self.sb.value() / 100.0

        try:
            self.map1.set_data(
                m0[y1:y2, x1:x2],
                cmap="Reds",
                title=self.selected_names[0],
                offset=(x1, y1),
            )
        except TypeError:
            self.map1.set_data(m0[y1:y2, x1:x2], cmap="Reds", title=self.selected_names[0])

        try:
            self.map2.set_data(
                m1[y1:y2, x1:x2],
                cmap="Greens",
                title=self.selected_names[1],
                offset=(x1, y1),
            )
        except TypeError:
            self.map2.set_data(m1[y1:y2, x1:x2], cmap="Greens", title=self.selected_names[1])
            """Nz norm."""

        try:
            self.rgb.set_data(
                m0[y1:y2, x1:x2],
                m1[y1:y2, x1:x2],
                m2[y1:y2, x1:x2],
                names=self.selected_symbols,
                br=br,
                bg=bg,
                bb=bb,
                offset=(x1, y1),
            )
        except TypeError:
            self.rgb.set_data(
                m0[y1:y2, x1:x2],
                m1[y1:y2, x1:x2],
                m2[y1:y2, x1:x2],
                names=self.selected_symbols,
                br=br,
                bg=bg,
                bb=bb,
            )

        self.img1.set_zoom(x1, y1, x2, y2, self.image_list[0][1])
        if len(self.image_list) > 1:
            self.img2.set_zoom(x1, y1, x2, y2, self.image_list[1][1])

        self.scatter.set_data(
            m0[y1:y2, x1:x2],
            m1[y1:y2, x1:x2],
            f"{self.selected_names[0]} Intensity",
            f"{self.selected_names[1]} Intensity",
            remember_limits=True,
        )

        # re-apply highlights respecting visibility
        self._apply_highlights_everywhere()
        self._clear_all_cursors()
        self._update_ternary_with_current_view()

    def _on_line_selected(self, line):
        self.latest_line = line
        if not self.selected_data or not self.selected_names or not self.rb_line.isChecked():
            return
        x1, y1, x2, y2 = line
        self._plot_line_scan(x1, y1, x2, y2)

    def _current_roi_slices(self):
        """Return (y1:y2, x1:x2) for current view.

        If no ROI, full image slice.
        """
        if not self.selected_data:
            return (slice(None), slice(None))
        m0 = self.selected_data[0]
        h, w = m0.shape[:2]
        if self.latest_roi:
            x1, y1, x2, y2 = self._fix_roi_bounds(self.latest_roi, w, h, min_size=2)
            return (slice(y1, y2), slice(x1, x2))
        return (slice(0, h), slice(0, w))

    def _update_ternary_with_current_view(self):
        if not (self._ternary_active and self._ternary_dlg and self.selected_data):
            return
        (ys, xs) = self._current_roi_slices()
        r, g, b = self.selected_data[:3]

        # current ROI arrays (flattened)
        a = r[ys, xs].astype(float).flatten()
        bb = g[ys, xs].astype(float).flatten()
        cc = b[ys, xs].astype(float).flatten()

        # set labels
        if getattr(self, "selected_symbols", None):
            self._ternary_dlg.set_labels(self.selected_symbols[:3])

        H = (ys.stop - ys.start) if isinstance(ys, slice) else len(np.arange(r.shape[0])[ys])
        W = (xs.stop - xs.start) if isinstance(xs, slice) else len(np.arange(r.shape[1])[xs])
        roi_shape = (H, W)
        roi_origin = (
            (xs.start if isinstance(xs, slice) else int(np.min(np.arange(r.shape[1])[xs]))),
            (ys.start if isinstance(ys, slice) else int(np.min(np.arange(r.shape[0])[ys]))),
        )

        self._ternary_dlg.update_points(a, bb, cc, roi_shape=roi_shape, roi_origin=roi_origin)

        # ensure ternary lasso follows your toggle (defaults True on first use)
        if not hasattr(self, "_ternary_lasso_enabled"):
            self._ternary_lasso_enabled = True
        self._ternary_dlg.enable_lasso(self._ternary_lasso_enabled)

    # ---------- Plots ----------
    def _plot_center_line(self):
        m0, m1, m2 = self.selected_data
        rows, cols = m0.shape
        y = rows // 2
        xs = np.arange(cols)
        p0 = (m0[y, :] / np.max(m0)) if np.max(m0) != 0 else m0[y, :]
        p1 = (m1[y, :] / np.max(m1)) if np.max(m1) != 0 else m1[y, :]
        p2 = (m2[y, :] / np.max(m2)) if np.max(m2) != 0 else m2[y, :]
        self.linescan.set_profiles(xs, [p0, p1, p2], self.selected_names)
        self._last_linescan = (xs, [p0, p1, p2], self.selected_names)
        self._line_coords = [(int(x), int(y)) for x in xs]

    def _plot_line_scan(self, x1, y1, x2, y2):
        def sample_line(img, x1, y1, x2, y2):
            n = int(max(abs(x2 - x1), abs(y2 - y1)))
            n = max(n, 2)
            xs = np.linspace(x1, x2, n).astype(int)
            ys = np.linspace(y1, y2, n).astype(int)
            xs = np.clip(xs, 0, img.shape[1] - 1)
            ys = np.clip(ys, 0, img.shape[0] - 1)
            return (
                np.arange(n),
                img[ys, xs],
                list(zip(xs.tolist(), ys.tolist(), strict=False)),
            )

        m0, m1, m2 = self.selected_data
        xs0, v0, coords0 = sample_line(m0, x1, y1, x2, y2)
        _, v1, _ = sample_line(m1, x1, y1, x2, y2)
        _, v2, _ = sample_line(m2, x1, y1, x2, y2)

        def nz_norm(v, base):
            mx = float(np.max(base))
            return (v / mx) if mx != 0 else v

        p0 = nz_norm(v0, m0)
        p1 = nz_norm(v1, m1)
        p2 = nz_norm(v2, m2)
        self.linescan.set_profiles(xs0, [p0, p1, p2], self.selected_names)
        self._last_linescan = (xs0, [p0, p1, p2], self.selected_names)
        self._line_coords = coords0

    # live pointer from line-scan hover
    def _on_linescan_hover(self, idx):
        if not self._line_coords or idx < 0 or idx >= len(self._line_coords):
            return
        x, y = self._line_coords[idx]
        self.map1.set_cursor_point(x, y)
        self.map2.set_cursor_point(x, y)
        self.rgb.set_cursor_point(x, y)
        self.img1.set_cursor_point(x, y)
        if len(self.image_list) > 1:
            self.img2.set_cursor_point(x, y)

    # ---------- Brightness ----------
    def _refresh_rgb(self, *_):
        if not self.selected_data:
            return
        r, g, b = self.selected_data
        self.rgb.refresh_brightness(
            self.sr.value() / 100,
            self.sg.value() / 100,
            self.sb.value() / 100,
            names=self.selected_names,
        )

    # ---------- Commands ----------
    def _save_roi(self):
        if not self.latest_roi:
            QMessageBox.warning(self, "Warning", "No ROI selected.")
            return
        export_dir = os.path.dirname(self.data_path)
        path = os.path.join(export_dir, "roi_coordinates.xlsx")
        try:
            if os.path.exists(path):
                wb = openpyxl.load_workbook(path)
                sh = wb.active
            else:
                wb = openpyxl.Workbook()
                sh = wb.active
                sh.append(["X1", "Y1", "X2", "Y2"])
            sh.append(list(self.latest_roi))
            wb.save(path)
            self.log.append(f"ROI saved to {path}.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save ROI: {e}")

    def _save_correlation(self):
        if not self.latest_roi or not self.selected_data:
            QMessageBox.warning(self, "Warning", "No ROI selected for correlation.")
            return
        x1, y1, x2, y2 = self.latest_roi
        x1, x2 = sorted((x1, x2))
        y1, y2 = sorted((y1, y2))
        d0 = self.selected_data[0][y1:y2, x1:x2].flatten()
        d1 = self.selected_data[1][y1:y2, x1:x2].flatten()
        export_dir = os.path.dirname(self.data_path)
        path = os.path.join(export_dir, "correlation_data.csv")
        try:
            with open(path, "w", newline="") as f:
                w = csv.writer(f)
                w.writerow(["Dataset 1", "Dataset 2"])
                w.writerows(zip(d0, d1, strict=False))
            self.log.append(f"Correlation data saved to {path}.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save correlation data: {e}")

    def _save_line_scan(self):
        """Set data path."""
        if not hasattr(self, "_last_linescan") or self._last_linescan is None:
            QMessageBox.warning(self, "Warning", "No line scan to save.")
            return
        xs, profiles, labels = self._last_linescan
        export_dir = os.path.dirname(self.data_path)
        path = os.path.join(export_dir, "line_scan.csv")
        try:
            with open(path, "w", newline="") as f:
                w = csv.writer(f)
                w.writerow(["Pixel"] + list(labels))
                for i in range(len(xs)):
                    row = [int(xs[i])]
                    for p in profiles:
                        row.append(float(p[i]))
                    w.writerow(row)
            self.log.append(f"Line scan saved to {path}.")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save line scan: {e}")

    def _manual_input(self):
        dlg = QDialog(self)
        dlg.setWindowTitle("Manual Input")
        lay = QVBoxLayout(dlg)
        lay.addWidget(QLabel("Enter coordinates (x1, y1, x2, y2):"))
        edit = QLineEdit()
        lay.addWidget(edit)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        lay.addWidget(btns)
        btns.accepted.connect(lambda: self._accept_manual(edit.text(), dlg))
        btns.rejected.connect(dlg.reject)
        dlg.exec_()

    def _accept_manual(self, text, dialog):
        try:
            x1, y1, x2, y2 = map(int, text.split(","))
            self.latest_roi = (x1, y1, x2, y2)
            self._on_roi_selected(self.latest_roi)
            dialog.accept()
        except Exception:
            QMessageBox.warning(
                self,
                "Invalid Input",
                "Please enter four integers separated by commas.",
            )

    def _quick_reset(self):
        if not self.selected_data or not self.selected_names:
            return

        r, g, b = self.selected_data
        names = self.selected_names
        br = self.sr.value() / 100.0
        bg = self.sg.value() / 100.0
        bb = self.sb.value() / 100.0

        # FULL VIEW -> origin reset
        self._roi_origin = (0, 0)

        try:
            self.map1.set_data(r, cmap="Reds", title=names[0], offset=(0, 0))
        except TypeError:
            self.map1.set_data(r, cmap="Reds", title=names[0])

        try:
            self.map2.set_data(g, cmap="Greens", title=names[1], offset=(0, 0))
        except TypeError:
            self.map2.set_data(g, cmap="Greens", title=names[1])

        try:
            self.rgb.set_data(
                r,
                g,
                b,
                names=self.selected_symbols,
                br=br,
                bg=bg,
                bb=bb,
                offset=(0, 0),
            )
        except TypeError:
            self.rgb.set_data(r, g, b, names=self.selected_symbols, br=br, bg=bg, bb=bb)

        if self.image_list:
            try:
                self.img1.set_data(self.image_list[0][1], offset=(0, 0))
            except TypeError:
                self.img1.set_data(self.image_list[0][1])
            if len(self.image_list) > 1:
                try:
                    self.img2.set_data(self.image_list[1][1], offset=(0, 0))
                except TypeError:
                    self.img2.set_data(self.image_list[1][1])

        if self.rb_corr.isChecked():
            self.scatter.set_data(
                r,
                g,
                f"{names[0]} Intensity",
                f"{names[1]} Intensity",
                remember_limits=False,
            )
        elif self.rb_line.isChecked():
            if getattr(self, "latest_line", None):
                x1, y1, x2, y2 = self.latest_line
                self._plot_line_scan(x1, y1, x2, y2)
            else:
                self._plot_center_line()

        # re-apply / respect visibility
        self._apply_highlights_everywhere()
        self._clear_all_cursors()
        self._update_ternary_with_current_view()

    def _full_reset(self):
        if self.rb_corr.isChecked():
            self._enter_correlation_mode(True)
        elif self.rb_line.isChecked():
            self._enter_linescan_mode(True)
        else:
            self._enter_pointdata_mode(True)

    def _wd(self, *parts):
        base = self.working_dir or os.getcwd()
        return os.path.abspath(os.path.join(base, *parts))

    def set_working_dir(self, folder: str):
        if not folder:
            return
        folder = os.path.abspath(folder)
        if folder == getattr(self, "working_dir", ""):
            return
        self.working_dir = folder
        fname = os.path.basename(getattr(self, "data_path", "") or "data.h5")
        self.set_data_path(os.path.join(folder, fname))

    def set_data_path(self, path: str):
        if not path:
            return
        new_path = os.path.abspath(path)
        if new_path == getattr(self, "data_path", ""):
            return
        self.data_path = new_path
        self.working_dir = os.path.dirname(new_path)

        # light restart of the DataLoader
        ld = getattr(self, "loader", None)
        if ld is not None:
            for m in ("requestInterruption", "quit"):
                if hasattr(ld, m):
                    try:
                        getattr(ld, m)()
                    except Exception:
                        pass
            if hasattr(ld, "wait"):
                try:
                    ld.wait(300)
                except Exception:
                    pass

        try:
            self.loader = DataLoader(data_path=self.data_path)
            self.loader.data_loaded.connect(self._on_data_loaded)
            self.loader.start()
        except Exception:
            pass

        if hasattr(self, "log"):
            try:
                self.log.append(f"Using data at: {self.data_path}")
            except Exception:
                pass


# Run standalone for testing
if __name__ == "__main__":
    import sys

    app = QApplication(sys.argv)
    w = AnalysisStep()
    w.setWindowTitle("AnalysisStep")
    w.show()
    sys.exit(app.exec_())
